"""
Run every SELECT query from _procurement, _financial, _leadership, _vendor
against PostgreSQL and write Query + Output to an Excel file.
"""
import sys
import json
import textwrap
from pathlib import Path
from datetime import datetime, date

sys.path.insert(0, str(Path(__file__).parent))

from database import get_connection, init_db
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Setup ─────────────────────────────────────────────────────────────────────

conn = get_connection()

# Build periods (same logic as kpi_engine)
ref_row = conn.execute(
    "SELECT MAX(document_date) FROM po_dump "
    "WHERE document_date IS NOT NULL AND document_date != ''"
).fetchone()
ref_date = ref_row[0] if (ref_row and ref_row[0]) else datetime.utcnow().strftime("%Y-%m-%d")
d = date.fromisoformat(ref_date)
fy_start_year = d.year if d.month >= 4 else d.year - 1
FY  = f"'{fy_start_year}-04-01'"
MTD = f"'{d.year}-{d.month:02d}-01'"

# high-value threshold
hv_row = conn.execute(
    "SELECT config_value FROM kpi_config WHERE config_key = 'HIGH_VALUE_PO_THRESHOLD'"
).fetchone()
HIGH_VAL = float(hv_row[0]) if hv_row else 10_000_000

# Reference date string for aging queries
ref_d_str = ref_date

# Company code config (blank = all)
cc_row = conn.execute(
    "SELECT config_value FROM kpi_config WHERE config_key = 'ACTIVE_COMPANY_CODES'"
).fetchone()
_cc_cfg = cc_row[0] if cc_row else ""
_cc_codes = [c.strip() for c in _cc_cfg.split(",") if c.strip()]
_CC_SQL = (
    "company_code IN (" + ",".join(f"'{c}'" for c in _cc_codes) + ")"
    if _cc_codes else "1=1"
)

# Cutoff for chart data (12 months back)
from datetime import timedelta
ref_d_obj  = date.fromisoformat(ref_date)
twelve_ago = (ref_d_obj.replace(day=1) - timedelta(days=335)).strftime("%Y-%m-%d")
CUTOFF = f"'{twelve_ago}'"

print(f"[config] ref_date={ref_date}  FY={FY}  MTD={MTD}  HIGH_VAL={HIGH_VAL:,.0f}")


# ── Query runner ──────────────────────────────────────────────────────────────

def run_q(label: str, sql: str) -> tuple[str, str]:
    """Execute sql, return (clean_sql, formatted_result_string)."""
    clean = textwrap.dedent(sql).strip()
    try:
        cur = conn.execute(sql)
        rows = cur.fetchall()
        if not rows:
            result = "NULL / No rows"
        elif len(rows) == 1 and len(rows[0]) == 1:
            # Single scalar
            val = rows[0][0]
            result = str(val) if val is not None else "NULL"
        else:
            # Multiple rows or columns — format as table
            lines = []
            for r in rows[:50]:          # cap at 50 rows for readability
                lines.append(" | ".join(str(v) for v in r))
            if len(rows) > 50:
                lines.append(f"... ({len(rows) - 50} more rows)")
            result = "\n".join(lines)
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        result = f"ERROR: {e}"
    print(f"  [{label}] done")
    return clean, result


# ── Collect all queries ───────────────────────────────────────────────────────

rows_data = []   # list of (section, kpi_code, query_sql, output)

def Q(section, code, label, sql):
    sql_clean, out = run_q(code, sql)
    rows_data.append((section, code, label, sql_clean, out))


# ═══════════════════════════════════════════════════════════════════════════════
print("\n[PROCUREMENT]")

Q("PROCUREMENT", "P1", "Total PO Value MTD", f"""
    SELECT SUM(CAST(net_order_value AS REAL))
    FROM po_dump
    WHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))
      AND document_date >= {MTD}
""")

Q("PROCUREMENT", "P2", "Active PO Count MTD", f"""
    SELECT COUNT(DISTINCT purchasing_document || '|' || item)
    FROM po_dump
    WHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))
      AND (delivery_completed IS NULL OR delivery_completed = '')
      AND document_date >= {MTD}
""")

Q("PROCUREMENT", "P3", f"High-Value PO Count (>{int(HIGH_VAL):,})", f"""
    SELECT COUNT(DISTINCT purchasing_document)
    FROM po_dump
    WHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))
      AND CAST(net_order_value AS REAL) > {HIGH_VAL}
""")

Q("PROCUREMENT", "P4", "Avg PR-to-PO Days", """
    SELECT AVG(CAST(pr_to_po_days AS REAL))
    FROM pr_po_grn_invoice
    WHERE pr_to_po_days IS NOT NULL
      AND pr_to_po_days >= 0
      AND purchase_requisition IS NOT NULL
      AND purchasing_document  IS NOT NULL
""")

Q("PROCUREMENT", "P5", "PO Approval Cycle (days)", """
    SELECT AVG((cl.change_date::DATE - po.document_date::DATE)::FLOAT)
    FROM po_dump po
    JOIN change_log cl
      ON cl.object_id    = po.purchasing_document
     AND cl.object_class = 'EINKBELEG'
     AND cl.field_name   IN ('FRGZU','FRGKE')
     AND cl.change_indicator = 'U'
     AND cl.new_value    = 'X'
    WHERE po.release_indicator = 'X'
""")

Q("PROCUREMENT", "P6", "PO Deletions MTD", f"""
    SELECT COUNT(DISTINCT purchasing_document)
    FROM po_dump
    WHERE deletion_indicator = 'L'
      AND document_date >= {MTD}
""")

Q("PROCUREMENT", "P7_amended", "PO Amendment Count", """
    SELECT COUNT(DISTINCT cl.object_id)
    FROM change_log cl
    WHERE cl.object_class    = 'EINKBELEG'
      AND cl.change_indicator = 'U'
      AND cl.field_name NOT IN ('FRGZU','FRGKE')
""")

Q("PROCUREMENT", "P7_total", "Total Active PO Count", """
    SELECT COUNT(DISTINCT purchasing_document) FROM po_dump
    WHERE deletion_indicator IS NULL OR deletion_indicator = ''
""")

Q("PROCUREMENT", "P8", "Open PR Lines >7 Days", f"""
    SELECT COUNT(DISTINCT pr.purchase_requisition || '|' || pr.item_of_requisition)
    FROM pr_dump pr
    WHERE pr.release_status IN ('X','XX','XXX','XXXX','XXXXX')
      AND (pr.deletion_indicator IS NULL OR pr.deletion_indicator = '')
      AND pr.release_date IS NOT NULL
      AND ('{ref_d_str}'::DATE - pr.release_date::DATE) > 7
      AND NOT EXISTS (
          SELECT 1 FROM po_dump po
          WHERE po.purchase_requisition = pr.purchase_requisition
            AND po.item_of_requisition  = pr.item_of_requisition
      )
""")

Q("PROCUREMENT", "P9", "Total PO Value YTD", f"""
    SELECT SUM(CAST(net_order_value AS REAL))
    FROM po_dump
    WHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))
      AND document_date >= {FY}
""")

Q("PROCUREMENT", "P10", "PO Line Count YTD", f"""
    SELECT COUNT(*) FROM po_dump
    WHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))
      AND document_date >= {FY}
""")

Q("PROCUREMENT", "AVG_PO_VALUE", "Average PO Value MTD", f"""
    SELECT AVG(CAST(net_order_value AS REAL))
    FROM po_dump
    WHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))
      AND document_date >= {MTD}
""")

Q("PROCUREMENT", "HV_TOTAL", "Total PO Count YTD (for HV rate)", f"""
    SELECT COUNT(DISTINCT purchasing_document) FROM po_dump
    WHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))
      AND document_date >= {FY}
""")

Q("PROCUREMENT", "MAVERICK_COUNT", "Maverick PO Count", """
    SELECT COUNT(DISTINCT purchasing_document) FROM po_dump
    WHERE (purchase_requisition IS NULL OR purchase_requisition = '')
      AND (deletion_indicator IS NULL OR deletion_indicator = '')
""")

Q("PROCUREMENT", "ACTIVE_VENDOR_YTD", "Active Vendors YTD", f"""
    SELECT COUNT(DISTINCT vendor) FROM po_dump
    WHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))
      AND document_date >= {FY}
""")


# ═══════════════════════════════════════════════════════════════════════════════
print("\n[FINANCIAL]")

Q("FINANCIAL", "F1", "Total Spend YTD (Invoices)", f"""
    WITH cancelled AS (
        SELECT DISTINCT reverse_invoice AS doc
        FROM   invoice_dump
        WHERE  reverse_invoice IS NOT NULL AND reverse_invoice != ''
    ),
    net_nonzero AS (
        SELECT company_code || '|' || vendor || '|' || vendor_invoice_ref AS grp
        FROM   invoice_dump
        WHERE  vendor_invoice_ref IS NOT NULL AND vendor_invoice_ref != ''
        GROUP  BY company_code, vendor, vendor_invoice_ref
        HAVING ABS(SUM(CAST(amount_local_ccy AS REAL))) > 0.005
    )
    SELECT SUM(
        CAST(amount_local_ccy AS REAL)
        * CASE WHEN debit_credit_ind = 'S' THEN 1.0 ELSE -1.0 END
    )
    FROM   invoice_dump
    WHERE  document_type IN ('RE','KR','RN')
      AND  posting_date >= {FY}
      AND  {_CC_SQL}
      AND  invoice_doc NOT IN (SELECT doc FROM cancelled)
      AND  (reverse_invoice IS NULL OR reverse_invoice = '')
      AND  (
           vendor_invoice_ref IS NULL OR vendor_invoice_ref = ''
           OR company_code || '|' || vendor || '|' || vendor_invoice_ref
              IN (SELECT grp FROM net_nonzero)
      )
""")

Q("FINANCIAL", "F2", "Invoice Cancellation Rate %", f"""
    SELECT
        COUNT(DISTINCT CASE WHEN document_type = 'RN' THEN invoice_doc END) * 100.0
        / NULLIF(COUNT(DISTINCT invoice_doc), 0)
    FROM invoice_dump
    WHERE {_CC_SQL}
""")

Q("FINANCIAL", "F3", "3-Way Match Success Rate %", f"""
    WITH po AS (
        SELECT company_code, purchasing_document, item,
               CAST(order_quantity  AS REAL) AS po_qty,
               CAST(net_order_value AS REAL) AS po_amt,
               CASE WHEN material_type IS NOT NULL AND material_type != ''
                    THEN 'MATERIAL' ELSE 'SERVICE' END AS item_type
        FROM po_dump
        WHERE {_CC_SQL}
          AND (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))
    ),
    grn AS (
        SELECT purchasing_document, item,
               SUM(CAST(quantity AS REAL)
                   * CASE WHEN debit_credit_ind='S' THEN 1.0 ELSE -1.0 END) AS grn_qty,
               SUM(CAST(amount_local_ccy AS REAL)
                   * CASE WHEN debit_credit_ind='S' THEN 1.0 ELSE -1.0 END) AS grn_amt
        FROM grn_dump GROUP BY purchasing_document, item
    ),
    inv AS (
        SELECT purchasing_document, item,
               SUM(CAST(quantity AS REAL)
                   * CASE WHEN debit_credit_ind='S' THEN 1.0 ELSE -1.0 END) AS inv_qty,
               SUM(CAST(amount_local_ccy AS REAL)
                   * CASE WHEN debit_credit_ind='S' THEN 1.0 ELSE -1.0 END) AS inv_amt
        FROM po_invoice_dump GROUP BY purchasing_document, item
    ),
    combined AS (
        SELECT po.item_type, po.po_qty, po.po_amt,
               COALESCE(grn.grn_qty,0) AS grn_qty, COALESCE(grn.grn_amt,0) AS grn_amt,
               COALESCE(inv.inv_qty,0) AS inv_qty, COALESCE(inv.inv_amt,0) AS inv_amt
        FROM po
        JOIN inv ON inv.purchasing_document=po.purchasing_document AND inv.item=po.item
        LEFT JOIN grn ON grn.purchasing_document=po.purchasing_document AND grn.item=po.item
    )
    SELECT
        COUNT(CASE
            WHEN item_type='MATERIAL'
              AND ABS(grn_qty-po_qty)/NULLIF(ABS(po_qty),0)<=0.05
              AND ABS(inv_qty-po_qty)/NULLIF(ABS(po_qty),0)<=0.05
              AND ABS(inv_qty-grn_qty)/NULLIF(ABS(grn_qty),0)<=0.05
            THEN 1
            WHEN item_type='SERVICE'
              AND ABS(grn_amt-po_amt)/NULLIF(ABS(po_amt),0)<=0.05
              AND ABS(inv_amt-po_amt)/NULLIF(ABS(po_amt),0)<=0.05
              AND ABS(inv_amt-grn_amt)/NULLIF(ABS(grn_amt),0)<=0.05
            THEN 1
        END)*100.0 / NULLIF(COUNT(*),0)
    FROM combined
""")

Q("FINANCIAL", "F4", "Invoice Processing Time (days)", f"""
    WITH cancelled_inv AS (
        SELECT DISTINCT reverse_invoice AS doc FROM invoice_dump
        WHERE reverse_invoice IS NOT NULL AND reverse_invoice != ''
    ),
    valid_inv AS (
        SELECT company_code, vendor, invoice_doc, vendor_invoice_date
        FROM invoice_dump
        WHERE {_CC_SQL}
          AND vendor_invoice_date IS NOT NULL
          AND (reverse_invoice IS NULL OR reverse_invoice = '')
          AND invoice_doc NOT IN (SELECT doc FROM cancelled_inv)
    ),
    noncancelled_pay AS (
        SELECT company_code, vendor, payment_doc
        FROM payment_dump
        GROUP BY company_code, vendor, payment_doc
        HAVING ABS(SUM(CAST(amount_local_ccy AS REAL)
            * CASE WHEN debit_credit_ind='S' THEN 1.0 ELSE -1.0 END)) > 0.005
    ),
    valid_pay AS (
        SELECT pd.company_code, pd.vendor, pd.clearing_date, pd.cleared_invoice
        FROM payment_dump pd
        JOIN noncancelled_pay nc ON nc.payment_doc=pd.payment_doc
          AND nc.company_code=pd.company_code AND nc.vendor=pd.vendor
        WHERE pd.debit_credit_ind='S'
    )
    SELECT AVG((vp.clearing_date::DATE - vi.vendor_invoice_date::DATE)::FLOAT)
    FROM valid_inv vi
    JOIN valid_pay vp ON vp.cleared_invoice=vi.invoice_doc
      AND vp.company_code=vi.company_code AND vp.vendor=vi.vendor
    WHERE vp.clearing_date IS NOT NULL
""")

Q("FINANCIAL", "F5", "On-Time Payment Rate %", f"""
    WITH cancelled_inv AS (
        SELECT DISTINCT reverse_invoice AS doc FROM invoice_dump
        WHERE reverse_invoice IS NOT NULL AND reverse_invoice != ''
    ),
    valid_inv AS (
        SELECT company_code, vendor, invoice_doc, due_date
        FROM invoice_dump
        WHERE {_CC_SQL} AND due_date IS NOT NULL AND due_date != ''
          AND (reverse_invoice IS NULL OR reverse_invoice = '')
          AND invoice_doc NOT IN (SELECT doc FROM cancelled_inv)
    ),
    noncancelled_pay AS (
        SELECT company_code, vendor, payment_doc
        FROM payment_dump
        GROUP BY company_code, vendor, payment_doc
        HAVING ABS(SUM(CAST(amount_local_ccy AS REAL)
            * CASE WHEN debit_credit_ind='S' THEN 1.0 ELSE -1.0 END)) > 0.005
    ),
    valid_pay AS (
        SELECT pd.company_code, pd.vendor, pd.clearing_date, pd.cleared_invoice
        FROM payment_dump pd
        JOIN noncancelled_pay nc ON nc.payment_doc=pd.payment_doc
          AND nc.company_code=pd.company_code AND nc.vendor=pd.vendor
        WHERE pd.debit_credit_ind='S'
    )
    SELECT COUNT(CASE WHEN vp.clearing_date=vi.due_date THEN 1 END)*100.0
           / NULLIF(COUNT(*),0)
    FROM valid_inv vi
    JOIN valid_pay vp ON vp.cleared_invoice=vi.invoice_doc
      AND vp.company_code=vi.company_code AND vp.vendor=vi.vendor
    WHERE vp.clearing_date IS NOT NULL
""")

Q("FINANCIAL", "F7", "Open Invoice Aging (total open amount)", f"""
    WITH open_inv AS (
        SELECT
            CAST(amount_local_ccy AS REAL)
                * CASE WHEN debit_credit_ind='S' THEN 1.0 ELSE -1.0 END AS signed_amt,
            (CURRENT_DATE - due_date::DATE) AS days_overdue
        FROM invoice_dump
        WHERE (clearing_doc IS NULL OR clearing_doc='')
          AND {_CC_SQL}
          AND document_type IN ('RE','KR')
          AND due_date IS NOT NULL AND due_date != ''
          AND (reverse_invoice IS NULL OR reverse_invoice='')
          AND invoice_doc NOT IN (
              SELECT DISTINCT reverse_invoice FROM invoice_dump
              WHERE reverse_invoice IS NOT NULL AND reverse_invoice != ''
          )
    )
    SELECT
        SUM(CASE WHEN days_overdue <   0 THEN signed_amt ELSE 0 END) AS not_yet_due,
        SUM(CASE WHEN days_overdue BETWEEN 0  AND  9 THEN signed_amt ELSE 0 END) AS d0_10,
        SUM(CASE WHEN days_overdue BETWEEN 10 AND 19 THEN signed_amt ELSE 0 END) AS d10_20,
        SUM(CASE WHEN days_overdue BETWEEN 20 AND 29 THEN signed_amt ELSE 0 END) AS d20_30,
        SUM(CASE WHEN days_overdue BETWEEN 30 AND 59 THEN signed_amt ELSE 0 END) AS d30_60,
        SUM(CASE WHEN days_overdue BETWEEN 60 AND 89 THEN signed_amt ELSE 0 END) AS d60_90,
        SUM(CASE WHEN days_overdue >= 90             THEN signed_amt ELSE 0 END) AS d90plus,
        SUM(signed_amt) AS total_open
    FROM open_inv
""")

Q("FINANCIAL", "F8", "Payment Timing (Early/OnTime/Late counts)", f"""
    WITH cancelled_inv AS (
        SELECT DISTINCT reverse_invoice AS doc FROM invoice_dump
        WHERE reverse_invoice IS NOT NULL AND reverse_invoice != ''
    ),
    valid_inv AS (
        SELECT company_code, vendor, invoice_doc, due_date
        FROM invoice_dump
        WHERE {_CC_SQL} AND due_date IS NOT NULL AND due_date != ''
          AND (reverse_invoice IS NULL OR reverse_invoice='')
          AND invoice_doc NOT IN (SELECT doc FROM cancelled_inv)
    ),
    noncancelled_pay AS (
        SELECT company_code, vendor, payment_doc
        FROM payment_dump GROUP BY company_code, vendor, payment_doc
        HAVING ABS(SUM(CAST(amount_local_ccy AS REAL)
            * CASE WHEN debit_credit_ind='S' THEN 1.0 ELSE -1.0 END)) > 0.005
    ),
    valid_pay AS (
        SELECT pd.company_code, pd.vendor, pd.clearing_date, pd.cleared_invoice
        FROM payment_dump pd
        JOIN noncancelled_pay nc ON nc.payment_doc=pd.payment_doc
          AND nc.company_code=pd.company_code AND nc.vendor=pd.vendor
        WHERE pd.debit_credit_ind='S'
    ),
    joined AS (
        SELECT (vp.clearing_date::DATE - vi.due_date::DATE) AS day_diff
        FROM valid_inv vi
        JOIN valid_pay vp ON vp.cleared_invoice=vi.invoice_doc
          AND vp.company_code=vi.company_code AND vp.vendor=vi.vendor
        WHERE vp.clearing_date IS NOT NULL
    )
    SELECT
        COUNT(CASE WHEN day_diff < 0 THEN 1 END) AS early,
        COUNT(CASE WHEN day_diff = 0 THEN 1 END) AS on_time,
        COUNT(CASE WHEN day_diff > 0 THEN 1 END) AS late,
        AVG(CASE WHEN day_diff < 0 THEN ABS(day_diff) END) AS avg_days_early,
        AVG(CASE WHEN day_diff > 0 THEN day_diff END) AS avg_days_late,
        COUNT(*) AS total
    FROM joined
""")

Q("FINANCIAL", "F10", "Total Payments YTD", f"""
    SELECT SUM(CAST(amount_local_ccy AS REAL))
    FROM payment_dump
    WHERE posting_date >= {FY}
""")

Q("FINANCIAL", "F11_PO_VAL", "Total PO Value YTD (for Payment/PO Ratio)", f"""
    SELECT SUM(CAST(net_order_value AS REAL))
    FROM po_dump
    WHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))
      AND document_date >= {FY}
""")

Q("FINANCIAL", "F_PR", "Approved PR Count YTD", f"""
    SELECT COUNT(DISTINCT purchase_requisition || '|' || item_of_requisition)
    FROM pr_dump
    WHERE release_status IN ('X','XX','XXX','XXXX','XXXXX')
      AND (deletion_indicator IS NULL OR deletion_indicator='')
      AND release_date >= {FY}
      AND {_CC_SQL}
""")


# ═══════════════════════════════════════════════════════════════════════════════
print("\n[LEADERSHIP]")

Q("LEADERSHIP", "L1", "Total PO Count & Value YTD", f"""
    SELECT
        SUM(CAST(net_order_value AS REAL)),
        COUNT(DISTINCT purchasing_document),
        COUNT(DISTINCT CASE WHEN release_indicator LIKE 'X%' THEN purchasing_document END),
        COUNT(DISTINCT CASE WHEN (release_indicator IS NULL OR release_indicator='')
                            THEN purchasing_document END)
    FROM po_dump
    WHERE (deletion_indicator IS NULL OR deletion_indicator <> 'L')
      AND created_on >= {FY}
      AND created_on <= CURRENT_DATE::TEXT
""")

Q("LEADERSHIP", "L_GRN", "GRN Count & Value YTD (mv_type=101)", f"""
    SELECT
        COUNT(DISTINCT material_document),
        SUM(CAST(amount_local_ccy AS REAL)
            * CASE WHEN debit_credit_ind='S' THEN 1.0 ELSE -1.0 END)
    FROM grn_dump
    WHERE movement_type='101'
      AND posting_date >= {FY}
""")

Q("LEADERSHIP", "L_INV", "Invoice Count & Value YTD (net of reversals)", f"""
    WITH cancelled AS (
        SELECT DISTINCT reverse_invoice AS doc FROM invoice_dump
        WHERE reverse_invoice IS NOT NULL AND reverse_invoice != ''
    )
    SELECT
        COUNT(DISTINCT invoice_doc),
        SUM(CAST(amount_local_ccy AS REAL)
            * CASE WHEN debit_credit_ind='S' THEN 1.0 ELSE -1.0 END)
    FROM invoice_dump
    WHERE document_type IN ('RE','KR')
      AND posting_date >= {FY}
      AND invoice_doc NOT IN (SELECT doc FROM cancelled)
      AND (reverse_invoice IS NULL OR reverse_invoice='')
""")

Q("LEADERSHIP", "L_INV_TYPE", "Invoice Summary by Vendor Type", """
    SELECT i.company_code,
           COALESCE(v.vendor_type,'UNKNOWN') AS vendor_type,
           COUNT(DISTINCT i.invoice_doc) AS invoice_count,
           SUM(CAST(i.amount_local_ccy AS REAL)
               * CASE WHEN i.debit_credit_ind='S' THEN 1.0 ELSE -1.0 END) AS total_amount
    FROM invoice_dump i
    LEFT JOIN vendor_master v ON i.vendor=v.vendor
    WHERE i.document_type IN ('RE','KR')
      AND (i.reverse_invoice IS NULL OR i.reverse_invoice='')
      AND i.invoice_doc NOT IN (
          SELECT DISTINCT reverse_invoice FROM invoice_dump
          WHERE reverse_invoice IS NOT NULL AND reverse_invoice != ''
      )
    GROUP BY i.company_code, v.vendor_type
    ORDER BY i.company_code, total_amount DESC
""")

Q("LEADERSHIP", "L2", "Maverick PO Rate %", """
    SELECT COUNT(CASE WHEN is_maverick=1 THEN 1 END)*100.0
           / NULLIF(COUNT(*),0)
    FROM pr_po_grn_invoice
    WHERE purchasing_document IS NOT NULL
""")

Q("LEADERSHIP", "L3", "E2E P2P Cycle Time (days)", """
    SELECT AVG(CAST(total_cycle_days AS REAL))
    FROM pr_po_grn_invoice
    WHERE total_cycle_days IS NOT NULL AND total_cycle_days > 0
""")

Q("LEADERSHIP", "L4_TOTAL", "Total PO Spend (for concentration)", """
    SELECT SUM(CAST(net_order_value AS REAL)) FROM po_dump
    WHERE (deletion_indicator IS NULL OR deletion_indicator <> 'L')
""")

Q("LEADERSHIP", "L4_TOP10", "Top-10 Vendors by Spend", """
    SELECT po.vendor,
           COALESCE(MIN(vm.vendor_name), MIN(po.vendor_name), po.vendor) AS name,
           SUM(CAST(po.net_order_value AS REAL)) AS spend
    FROM po_dump po
    LEFT JOIN vendor_master vm ON po.vendor=vm.vendor
    WHERE (po.deletion_indicator IS NULL OR po.deletion_indicator <> 'L')
    GROUP BY po.vendor
    ORDER BY spend DESC
    LIMIT 10
""")

Q("LEADERSHIP", "L4b_SS_TOTAL", "Single Source Procurement Count & Value", """
    SELECT COUNT(*), SUM(total_value)
    FROM (
        SELECT company_code, material_description,
               SUM(CAST(net_order_value AS REAL)) AS total_value
        FROM po_dump
        WHERE (deletion_indicator IS NULL OR deletion_indicator <> 'L')
          AND material_description IS NOT NULL AND material_description != ''
        GROUP BY company_code, material_description
        HAVING COUNT(DISTINCT vendor) = 1
    ) t
""")

Q("LEADERSHIP", "L4b_SS_TOP20", "Top-20 Single Source Items", """
    WITH ss AS (
        SELECT po.company_code, po.material_description,
               MIN(po.vendor) AS vendor_code,
               SUM(CAST(po.net_order_value AS REAL)) AS total_value
        FROM po_dump po
        WHERE (po.deletion_indicator IS NULL OR po.deletion_indicator <> 'L')
          AND po.material_description IS NOT NULL AND po.material_description != ''
        GROUP BY po.company_code, po.material_description
        HAVING COUNT(DISTINCT po.vendor)=1
    )
    SELECT ss.company_code, ss.material_description, ss.vendor_code,
           COALESCE(vm.vendor_name, ss.vendor_code) AS vendor_name,
           ROUND(ss.total_value,2)
    FROM ss
    LEFT JOIN vendor_master vm ON ss.vendor_code=vm.vendor
    ORDER BY ss.total_value DESC
    LIMIT 20
""")

Q("LEADERSHIP", "L5", "Negotiation Savings YTD", """
    SELECT SUM((CAST(f.pr_value AS REAL) - CAST(f.po_net_price AS REAL))
               * CAST(f.po_quantity AS REAL))
    FROM pr_po_grn_invoice f
    WHERE f.pr_value IS NOT NULL AND f.po_net_price IS NOT NULL
      AND (CAST(f.pr_value AS REAL) - CAST(f.po_net_price AS REAL)) > 0
""")

Q("LEADERSHIP", "L6_ANOM", "Anomaly Count (for Risk Score)", """
    SELECT COUNT(*) FROM process_mining_events WHERE anomaly_count > 0
""")

Q("LEADERSHIP", "L6_TOTAL_EV", "Total Events (for Risk Score)", """
    SELECT COUNT(*) FROM process_mining_events
""")

Q("LEADERSHIP", "S7a", "SOD: PO Create vs Release", """
    SELECT COUNT(DISTINCT po.purchasing_document)
    FROM po_dump po
    JOIN change_log cl ON cl.object_id=po.purchasing_document
    WHERE cl.object_class='EINKBELEG'
      AND cl.table_name='EKKO'
      AND cl.field_name='FRGZU'
      AND cl.change_indicator IN ('E','U')
      AND cl.new_value='X'
      AND cl.username=po.created_by
      AND (po.deletion_indicator IS NULL OR po.deletion_indicator='')
""")

Q("LEADERSHIP", "S7b", "SOD: PO Create vs GRN", """
    SELECT COUNT(DISTINCT po.purchasing_document || '|' || po.item)
    FROM po_dump po
    JOIN grn_dump grn ON grn.purchasing_document=po.purchasing_document
                     AND grn.item=po.item
    WHERE po.created_by=grn.created_by
      AND (po.deletion_indicator IS NULL OR po.deletion_indicator='')
      AND grn.debit_credit_ind='S'
""")

Q("LEADERSHIP", "S7c", "SOD: GRN vs Invoice", """
    SELECT COUNT(DISTINCT grn.material_document)
    FROM grn_dump grn
    JOIN po_invoice_dump inv ON inv.purchasing_document=grn.purchasing_document
                            AND inv.item=grn.item
    WHERE grn.created_by=inv.created_by
      AND grn.debit_credit_ind='S'
      AND inv.debit_credit_ind='S'
""")

Q("LEADERSHIP", "S7d", "SOD: Invoice vs Payment", """
    SELECT COUNT(DISTINCT i.invoice_doc)
    FROM invoice_dump i
    JOIN payment_dump p ON p.company_code=i.company_code
                       AND p.vendor=i.vendor
                       AND p.cleared_invoice=i.invoice_doc
    WHERE i.created_by=p.created_by
      AND i.document_type IN ('RE','KR')
      AND (i.reverse_invoice IS NULL OR i.reverse_invoice='')
      AND i.invoice_doc NOT IN (
          SELECT DISTINCT reverse_invoice FROM invoice_dump
          WHERE reverse_invoice IS NOT NULL AND reverse_invoice != ''
      )
      AND p.debit_credit_ind='S'
""")

Q("LEADERSHIP", "L8_TOTAL", "PO Lines in Fact Table", """
    SELECT COUNT(DISTINCT purchasing_document || '|' || item)
    FROM pr_po_grn_invoice WHERE purchasing_document IS NOT NULL
""")

Q("LEADERSHIP", "L8_WITH_GRN", "PO Lines with GRN", """
    SELECT COUNT(DISTINCT purchasing_document || '|' || item)
    FROM pr_po_grn_invoice
    WHERE purchasing_document IS NOT NULL AND grn_posting_date IS NOT NULL
""")

Q("LEADERSHIP", "L9", "Duplicate Invoice Count", """
    SELECT SUM(cnt-1) FROM (
        SELECT COUNT(DISTINCT invoice_doc) AS cnt
        FROM invoice_dump
        WHERE document_type IN ('RE','KR')
          AND (reverse_invoice IS NULL OR reverse_invoice='')
          AND invoice_doc NOT IN (
              SELECT DISTINCT reverse_invoice FROM invoice_dump
              WHERE reverse_invoice IS NOT NULL AND reverse_invoice != ''
          )
        GROUP BY company_code, vendor_invoice_ref,
                 CAST(amount_local_ccy AS REAL),
                 vendor, posting_date
        HAVING COUNT(DISTINCT invoice_doc) > 1
    ) t
""")

Q("LEADERSHIP", "L9b", "Duplicate PO Count", """
    SELECT SUM(cnt-1) FROM (
        SELECT COUNT(DISTINCT purchasing_document) AS cnt
        FROM po_dump
        WHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))
        GROUP BY company_code, material_group, vendor,
                 CAST(net_order_value AS REAL),
                 CAST(order_quantity AS REAL),
                 document_date
        HAVING COUNT(DISTINCT purchasing_document) > 1
    ) t
""")

Q("LEADERSHIP", "L10", f"High-Value PO Count (>{int(HIGH_VAL):,})", f"""
    SELECT COUNT(DISTINCT purchasing_document)
    FROM po_dump
    WHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))
      AND CAST(net_order_value AS REAL) > {HIGH_VAL}
""")

Q("LEADERSHIP", "L11a", "PR Amount YTD", f"""
    SELECT SUM(CAST(valuation_price AS REAL))
    FROM pr_dump
    WHERE (deletion_indicator IS NULL OR deletion_indicator='')
      AND release_date >= {FY}
""")

Q("LEADERSHIP", "L11b", "PR Line Count YTD", f"""
    SELECT COUNT(DISTINCT purchase_requisition || '|' || item_of_requisition)
    FROM pr_dump
    WHERE (deletion_indicator IS NULL OR deletion_indicator='')
      AND release_date >= {FY}
""")

Q("LEADERSHIP", "L11c", "PO Line Count YTD", f"""
    SELECT COUNT(*) FROM po_dump
    WHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))
      AND document_date >= {FY}
""")

Q("LEADERSHIP", "L11d", "One-Time Vendor Count", """
    SELECT COUNT(*) FROM vendor_master WHERE UPPER(vendor_type)='ONE_TIME'
""")

Q("LEADERSHIP", "L11e", "PO Lines without Contract", """
    SELECT COUNT(*) FROM po_dump
    WHERE (contract_number IS NULL OR contract_number='')
      AND (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))
""")

Q("LEADERSHIP", "L12_SUMMARY", "P2P Summary Counts", """
    SELECT
        COUNT(DISTINCT purchase_requisition || '|' || item_of_requisition) AS approved_pr
    FROM pr_dump WHERE release_status IN ('X','XX','XXX','XXXX','XXXXX')
""")


# ═══════════════════════════════════════════════════════════════════════════════
print("\n[VENDOR]")

Q("VENDOR", "V1", "Active Vendor Count", """
    SELECT COUNT(DISTINCT vendor)
    FROM vendor_master
    WHERE (central_purchasing_block IS NULL OR central_purchasing_block <> 'X')
      AND (central_posting_block    IS NULL OR central_posting_block    <> 'X')
      AND (deletion_flag_central    IS NULL OR deletion_flag_central    <> 'X')
      AND (payment_block            IS NULL OR payment_block            <> '*')
      AND (posting_block_cc         IS NULL OR posting_block_cc         <> 'X')
""")

Q("VENDOR", "V2", "Vendor Health Breakdown", """
    SELECT
        COUNT(CASE WHEN (central_purchasing_block IS NULL OR central_purchasing_block<>'X')
                     AND (payment_block IS NULL OR payment_block<>'*')
                     AND (posting_block_cc IS NULL OR posting_block_cc<>'X')
                   THEN 1 END) AS active,
        COUNT(CASE WHEN central_purchasing_block='X'
                      OR payment_block='*'
                      OR posting_block_cc='X'
                   THEN 1 END) AS non_active,
        COUNT(CASE WHEN UPPER(vendor_type)='ONE_TIME'      THEN 1 END) AS one_time,
        COUNT(CASE WHEN UPPER(vendor_type)='DOMESTIC'      THEN 1 END) AS domestic,
        COUNT(CASE WHEN UPPER(vendor_type)='INTERNATIONAL' THEN 1 END) AS international,
        COUNT(CASE WHEN msme_flag IN ('M','S')             THEN 1 END) AS msme,
        COUNT(*) AS total
    FROM vendor_master
""")

Q("VENDOR", "V3_TOP25", "Vendor Delivery Lead Time (top 25 by avg days)", """
    WITH grn_first AS (
        SELECT purchasing_document, item,
               MIN(entry_date) AS first_grn_date
        FROM grn_dump
        WHERE debit_credit_ind='S'
          AND entry_date IS NOT NULL AND entry_date != ''
        GROUP BY purchasing_document, item
    )
    SELECT po.vendor,
           COALESCE(MIN(vm.vendor_name), MIN(po.vendor_name), po.vendor) AS vendor_name,
           ROUND(AVG((gf.first_grn_date::DATE - po.delivery_date::DATE)::FLOAT),1) AS avg_days,
           COUNT(*) AS po_lines
    FROM po_dump po
    JOIN grn_first gf ON gf.purchasing_document=po.purchasing_document
                     AND gf.item=po.item
    LEFT JOIN vendor_master vm ON po.vendor=vm.vendor
    WHERE (po.deletion_indicator IS NULL OR po.deletion_indicator NOT IN ('L','X'))
      AND po.delivery_date IS NOT NULL AND po.delivery_date != ''
    GROUP BY po.vendor
    ORDER BY avg_days DESC
    LIMIT 25
""")

Q("VENDOR", "V3_OVERALL", "Vendor Delivery Lead Time (overall avg)", """
    WITH grn_first AS (
        SELECT purchasing_document, item,
               MIN(entry_date) AS first_grn_date
        FROM grn_dump
        WHERE debit_credit_ind='S'
          AND entry_date IS NOT NULL AND entry_date != ''
        GROUP BY purchasing_document, item
    )
    SELECT ROUND(AVG((gf.first_grn_date::DATE - po.delivery_date::DATE)::FLOAT),1)
    FROM po_dump po
    JOIN grn_first gf ON gf.purchasing_document=po.purchasing_document
                     AND gf.item=po.item
    WHERE (po.deletion_indicator IS NULL OR po.deletion_indicator NOT IN ('L','X'))
      AND po.delivery_date IS NOT NULL AND po.delivery_date != ''
""")

Q("VENDOR", "V4", "Avg Delivery Delay (late only, days)", """
    SELECT AVG((grn.posting_date::DATE - pod.expected_delivery_date::DATE)::FLOAT)
    FROM po_delivery_dump pod
    JOIN grn_dump grn ON grn.purchasing_document=pod.purchasing_document
                     AND grn.item=pod.item
    WHERE grn.debit_credit_ind='S'
      AND grn.posting_date > pod.expected_delivery_date
""")

Q("VENDOR", "V6_TOP10", "Top-10 Vendors by Spend Share", """
    SELECT po.vendor,
           COALESCE(MIN(vm.vendor_name), MIN(po.vendor_name), po.vendor) AS name,
           SUM(CAST(po.net_order_value AS REAL)) AS spend
    FROM po_dump po
    LEFT JOIN vendor_master vm ON po.vendor=vm.vendor
    WHERE po.deletion_indicator IS NULL OR po.deletion_indicator NOT IN ('L','X')
    GROUP BY po.vendor
    ORDER BY spend DESC
    LIMIT 10
""")

Q("VENDOR", "V7", "Blocked Vendor Count", """
    SELECT COUNT(DISTINCT vendor) FROM vendor_master
    WHERE central_purchasing_block='X'
       OR central_posting_block='X'
       OR payment_block='*'
       OR posting_block_cc='X'
""")

Q("VENDOR", "V8", "MSME Vendor Count", """
    SELECT COUNT(*) FROM vendor_master WHERE msme_flag IN ('M','S')
""")

Q("VENDOR", "V9_CLEAR", "Vendors with All Blocks Clear", """
    SELECT COUNT(*) FROM vendor_master
    WHERE (central_purchasing_block IS NULL OR central_purchasing_block='')
      AND (central_posting_block    IS NULL OR central_posting_block   ='')
      AND (deletion_flag_central    IS NULL OR deletion_flag_central   ='')
      AND (payment_block            IS NULL OR payment_block           ='')
      AND (posting_block_cc         IS NULL OR posting_block_cc        ='')
""")

Q("VENDOR", "V9_TOTAL", "Total Vendor Count", """
    SELECT COUNT(*) FROM vendor_master
""")

Q("VENDOR", "V10", "Vendor Master Changes MTD", f"""
    SELECT COUNT(DISTINCT object_id)
    FROM change_log
    WHERE object_class='KRED'
      AND change_date >= {MTD}
""")


# ═══════════════════════════════════════════════════════════════════════════════
# Write Excel
# ═══════════════════════════════════════════════════════════════════════════════
print(f"\n[Excel] Writing {len(rows_data)} rows…")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "KPI Query Results"

# Colour palette
HDR_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
SEC_FILLS  = {
    "PROCUREMENT": PatternFill("solid", fgColor="2E75B6"),
    "FINANCIAL":   PatternFill("solid", fgColor="375623"),
    "LEADERSHIP":  PatternFill("solid", fgColor="7B2C2C"),
    "VENDOR":      PatternFill("solid", fgColor="6B4E8A"),
}
SEC_TEXT   = Font(color="FFFFFF", bold=True, size=11)
MONO_FONT  = Font(name="Courier New", size=9)
LABEL_FONT = Font(bold=True, size=10)
BORDER_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# Header row
headers = ["Section", "KPI Code", "KPI Label", "SQL Query", "Output / Result"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill  = HDR_FILL
    cell.font  = Font(color="FFFFFF", bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER_THIN

ws.row_dimensions[1].height = 22

# Data rows
current_section = None
for i, (section, code, label, sql_clean, output) in enumerate(rows_data, start=2):
    sec_fill = SEC_FILLS.get(section, PatternFill("solid", fgColor="DDDDDD"))

    # Section
    c = ws.cell(row=i, column=1, value=section)
    c.fill      = sec_fill
    c.font      = SEC_TEXT
    c.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    c.border    = BORDER_THIN

    # KPI Code
    c = ws.cell(row=i, column=2, value=code)
    c.font      = Font(bold=True, size=10, color="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="top")
    c.border    = BORDER_THIN

    # Label
    c = ws.cell(row=i, column=3, value=label)
    c.font      = LABEL_FONT
    c.alignment = Alignment(vertical="top", wrap_text=True)
    c.border    = BORDER_THIN

    # SQL Query
    c = ws.cell(row=i, column=4, value=sql_clean)
    c.font      = MONO_FONT
    c.alignment = Alignment(vertical="top", wrap_text=True)
    c.fill      = PatternFill("solid", fgColor="F2F2F2")
    c.border    = BORDER_THIN

    # Output
    c = ws.cell(row=i, column=5, value=output)
    is_error = output.startswith("ERROR")
    c.font      = Font(
        name="Courier New", size=9,
        color="CC0000" if is_error else "1A1A1A",
        bold=is_error,
    )
    c.alignment = Alignment(vertical="top", wrap_text=True)
    c.fill      = PatternFill("solid", fgColor="FFF2F2" if is_error else "FFFFFF")
    c.border    = BORDER_THIN

    # Row height
    lines = max(output.count("\n") + 1, sql_clean.count("\n") + 1, 3)
    ws.row_dimensions[i].height = min(lines * 14, 200)

# Column widths
ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 32
ws.column_dimensions["D"].width = 70
ws.column_dimensions["E"].width = 60

# Freeze header
ws.freeze_panes = "A2"

out_path = Path(__file__).parent.parent / "KPI_Query_Results.xlsx"
wb.save(out_path)
print(f"[Excel] Saved -> {out_path}")
print(f"[Done] {len(rows_data)} queries executed.")
