import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\satya\OneDrive\Desktop\IntelliflowV2\Copy of IntelliSource_P2P_Reference_Schema_2.xlsx')

data_sheets = ['01_PR_Dump','02_PO_Dump','03_PO_Delivery_Dump','04_GRN_Dump','05_PO_Invoice_Dump','06_Invoice_Dump','07_Payment_Dump','08_Vendor_Master','09_Change_Log']

for name in data_sheets:
    ws = wb[name]
    print(f'\n=== {name} ===')
    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    header_row_idx = None
    for idx, row in enumerate(all_rows):
        if row and len(row) > 1 and str(row[0]).strip() == '#' and str(row[1]).strip() == 'Field name':
            header_row_idx = idx
            break
    if header_row_idx is None:
        print('  Could not find header row')
        continue
    fname_idx = 1
    fields = []
    for row in all_rows[header_row_idx+1:]:
        if row and row[fname_idx] and str(row[fname_idx]).strip():
            field_name = str(row[fname_idx]).strip()
            if field_name.lower() == 'field name':
                continue
            fields.append(field_name)
    for i, f in enumerate(fields, 1):
        print(f'  {i}. {f}')
