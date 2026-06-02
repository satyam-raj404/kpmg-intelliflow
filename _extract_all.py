import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\satya\OneDrive\Desktop\IntelliflowV2\Copy of IntelliSource_P2P_Reference_Schema_2.xlsx')

for name in wb.sheetnames:
    ws = wb[name]
    print(f'\n{"="*60}')
    print(f'Sheet: {name}  (Rows: {ws.max_row}, Cols: {ws.max_column})')
    print(f'{"="*60}')
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)):
        vals = [str(c).strip() if c is not None else '' for c in row]
        # Only print non-empty rows
        if any(v for v in vals):
            print(f'Row {idx}: {vals}')
