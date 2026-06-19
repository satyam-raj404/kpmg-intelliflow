"""Generate Utilization_Schema_and_KPIs.xlsx in the repo root."""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── helpers ──────────────────────────────────────────────────────────────────
NAVY   = "1F3864"
BLUE   = "2F5496"
LBLUE  = "D6E4F7"
GREEN  = "375623"
LGREEN = "E2EFDA"
AMBER  = "7F6000"
LAMBER = "FFEB9C"
GREY   = "F2F2F2"
WHITE  = "FFFFFF"
RED    = "C00000"

def hdr(ws, row, col, val, bg=NAVY, fg=WHITE, bold=True, sz=11, wrap=False, halign="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=fg, size=sz)
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    return c

def cell(ws, row, col, val, bg=None, fg="000000", bold=False, wrap=False, halign="left", italic=False):
    c = ws.cell(row=row, column=col, value=val)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(color=fg, bold=bold, italic=italic)
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    return c

def border_range(ws, min_row, max_row, min_col, max_col, style="thin"):
    bd = Side(style=style)
    for r in range(min_row, max_row+1):
        for c in range(min_col, max_col+1):
            existing = ws.cell(r, c).border
            top    = bd if r == min_row else existing.top
            bottom = bd if r == max_row else existing.bottom
            left   = bd if c == min_col else existing.left
            right  = bd if c == max_col else existing.right
            ws.cell(r, c).border = Border(top=top, bottom=bottom, left=left, right=right)

def freeze(ws, row, col):
    ws.freeze_panes = ws.cell(row=row, column=col)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Overview
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Overview"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 40

hdr(ws1, 1, 1, "IntelliSource — Utilization Dashboard: Schema & KPI Reference", bg=NAVY, sz=14)
ws1.merge_cells("A1:E1")

rows = [
    (2,  "Document",      "Utilization_Schema_and_KPIs.xlsx"),
    (3,  "Project",       "IntelliSource P2P Intelligence Platform"),
    (4,  "Dashboard",     "Utilization (/utilization)"),
    (5,  "Backend",       "backend/services/kpi_engine.py → _utilization(), _utilization_software(), _utilization_materials()"),
    (6,  "Frontend",      "kpmg-intelliflow/src/routes/utilization.tsx"),
    (7,  "DB Engine",     "PostgreSQL (psycopg3) — kpi_results table, ON CONFLICT upsert"),
    (8,  "KPI Compute",   "compute_all() → per-company loop (1001, 1002, 1003) + ALL aggregate"),
    (9,  "Company Filter","Uses useKpiCompanies('utilization') hook; CompanyFilter renders only when >1 company in kpi_results"),
    (10, "Total KPIs",    "35 KPI codes × 4 company variants (1001/1002/1003/ALL) = 140 rows in kpi_results"),
    (11, "Sheets",        "Overview | New Tables | Schema Changes | Base KPIs | Software KPIs | Materials KPIs | Auto-Categorization | Data Flow"),
]
for r, label, val in rows:
    cell(ws1, r, 1, label, bg=LBLUE, bold=True, halign="right")
    cell(ws1, r, 2, val, wrap=True)
    ws1.merge_cells(f"B{r}:E{r}")
    ws1.row_dimensions[r].height = 20

set_col_widths(ws1, [22, 90, 10, 10, 10])


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — New Tables
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("New Tables")
ws2.sheet_view.showGridLines = False

hdr(ws2, 1, 1, "New Tables Added for Utilization Dashboard", bg=NAVY, sz=13)
ws2.merge_cells("A1:F1")

tables = [
    {
        "name": "po_categorization",
        "purpose": "Tags each PO line with category (SOFTWARE/MATERIAL/SERVICE), CAPEX/OPEX flag, profit center, license type. Supports both SYSTEM auto-tags and manual user overrides.",
        "cols": [
            ("purchasing_document", "TEXT", "PK part", "PO number"),
            ("item",                "TEXT", "PK part", "PO line item"),
            ("po_category",         "TEXT", "", "SOFTWARE / MATERIAL / SERVICE"),
            ("sub_category",        "TEXT", "", "e.g. SaaS, On-Premise, Raw Material"),
            ("capex_opex_flag",     "TEXT", "", "CAPEX or OPEX"),
            ("profit_center",       "TEXT", "", "Cost allocation centre"),
            ("license_type",        "TEXT", "", "SUBSCRIPTION / PERPETUAL / ROYALTY / …"),
            ("budget_ref",          "TEXT", "", "Reference to pc_budget"),
            ("tagged_by",           "TEXT", "", "SYSTEM (auto) or USER (manual)"),
            ("tagged_at",           "TIMESTAMP", "", "When tagged"),
            ("notes",               "TEXT", "", "Freeform notes"),
        ]
    },
    {
        "name": "material_license_cost",
        "purpose": "Stores licensing fee details for material PO lines (royalties, import licences, etc.).",
        "cols": [
            ("purchasing_document", "TEXT", "FK po_dump", "PO number"),
            ("item",                "TEXT", "FK po_dump", "PO line item"),
            ("license_type",        "TEXT", "", "ROYALTY / IMPORT_LIC / PATENT / …"),
            ("license_fee_inr",     "REAL", "", "Fee in INR"),
            ("fee_basis",           "TEXT", "", "FIXED / PCT_OF_PO / PER_UNIT"),
            ("fee_pct",             "REAL", "", "% if fee_basis=PCT_OF_PO"),
            ("fee_per_unit",        "REAL", "", "Amount per unit if fee_basis=PER_UNIT"),
            ("vendor",              "TEXT", "", "Licensor name"),
            ("validity_start",      "DATE", "", "Licence start date"),
            ("validity_end",        "DATE", "", "Licence end date"),
            ("created_by",          "TEXT", "", "Who created this record"),
            ("created_at",          "TIMESTAMP", "", "Creation timestamp"),
            ("notes",               "TEXT", "", "Freeform notes"),
        ]
    },
    {
        "name": "profit_center_master",
        "purpose": "Maps profit centres to companies and business units. Used to allocate PO spend to cost centres.",
        "cols": [
            ("profit_center",       "TEXT", "PK", "Profit centre code (e.g. PC1001-A)"),
            ("pc_name",             "TEXT", "", "Display name"),
            ("company_code",        "TEXT", "FK", "Company code (1001/1002/1003)"),
            ("responsible_person",  "TEXT", "", "Owner name"),
            ("bu_type",             "TEXT", "", "IT / Finance / Operations / R&D / …"),
            ("cost_center_range",   "TEXT", "", "SAP cost centre range"),
            ("is_active",           "BOOLEAN", "", "Whether PC is active"),
        ]
    },
    {
        "name": "pc_budget",
        "purpose": "Annual CAPEX/OPEX budget per profit centre. Used to compute budget utilisation %.",
        "cols": [
            ("profit_center",  "TEXT", "PK part", "Profit centre code"),
            ("fiscal_year",    "INTEGER", "PK part", "4-digit year"),
            ("budget_type",    "TEXT", "PK part", "CAPEX or OPEX"),
            ("budget_inr",     "REAL", "", "Approved budget in INR"),
            ("approved_by",    "TEXT", "", "Approver name"),
            ("approved_at",    "DATE", "", "Approval date"),
        ]
    },
]

row = 2
for tbl in tables:
    # Table header
    hdr(ws2, row, 1, tbl["name"], bg=BLUE, sz=12)
    ws2.merge_cells(f"A{row}:F{row}")
    row += 1

    cell(ws2, row, 1, "Purpose:", bold=True, bg=GREY)
    cell(ws2, row, 2, tbl["purpose"], wrap=True, bg=GREY)
    ws2.merge_cells(f"B{row}:F{row}")
    ws2.row_dimensions[row].height = 30
    row += 1

    # Column header
    for c, h in enumerate(["Column", "Type", "Key/FK", "Description"], 1):
        hdr(ws2, row, c, h, bg=LBLUE, fg="000000", sz=10)
    row += 1

    for cname, ctype, key, desc in tbl["cols"]:
        cell(ws2, row, 1, cname, bold=True)
        cell(ws2, row, 2, ctype, halign="center")
        cell(ws2, row, 3, key, halign="center", fg="C00000", bold=bool(key))
        cell(ws2, row, 4, desc, wrap=True)
        ws2.merge_cells(f"D{row}:F{row}")
        row += 1

    border_range(ws2, row - len(tbl["cols"]) - 1, row - 1, 1, 4)
    row += 1  # gap

set_col_widths(ws2, [28, 14, 14, 60, 10, 10])


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Schema Changes (license_usage extended columns)
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Schema Changes")
ws3.sheet_view.showGridLines = False

hdr(ws3, 1, 1, "Schema Changes — Existing Table Alterations", bg=NAVY, sz=13)
ws3.merge_cells("A1:G1")

hdr(ws3, 2, 1, "Table: license_usage — New Columns Added via ALTER TABLE", bg=BLUE)
ws3.merge_cells(f"A2:G2")

cols_hdr = ["Column Added", "Type", "Default", "Purpose", "Migration Command", "", ""]
for c, h in enumerate(cols_hdr[:5], 1):
    hdr(ws3, 3, c, h, bg=LBLUE, fg="000000", sz=10)

lu_cols = [
    ("profit_center",  "TEXT", "''",              "Links license to profit centre for cost allocation",
     "ALTER TABLE license_usage ADD COLUMN IF NOT EXISTS profit_center TEXT DEFAULT ''"),
    ("license_type",   "TEXT", "'SUBSCRIPTION'",  "Type of licence: SUBSCRIPTION / PERPETUAL / ENTERPRISE / …",
     "ALTER TABLE license_usage ADD COLUMN IF NOT EXISTS license_type TEXT DEFAULT 'SUBSCRIPTION'"),
    ("vendor",         "TEXT", "''",              "Vendor / licensor name (independent of po_dump vendor)",
     "ALTER TABLE license_usage ADD COLUMN IF NOT EXISTS vendor TEXT DEFAULT ''"),
    ("po_reference",   "TEXT", "''",              "Linked PO number for reconciliation with po_dump",
     "ALTER TABLE license_usage ADD COLUMN IF NOT EXISTS po_reference TEXT DEFAULT ''"),
]
for i, (col, typ, dflt, purpose, cmd) in enumerate(lu_cols, 4):
    bg = GREY if i % 2 == 0 else WHITE
    cell(ws3, i, 1, col, bold=True, bg=bg)
    cell(ws3, i, 2, typ, bg=bg, halign="center")
    cell(ws3, i, 3, dflt, bg=bg, halign="center", italic=True)
    cell(ws3, i, 4, purpose, bg=bg, wrap=True)
    cell(ws3, i, 5, cmd, bg=bg, wrap=True, italic=True, fg="444444")
    ws3.merge_cells(f"E{i}:G{i}")
    ws3.row_dimensions[i].height = 22

border_range(ws3, 3, 7, 1, 5)

# Migration notes
r = 9
hdr(ws3, r, 1, "Migration Notes", bg=BLUE)
ws3.merge_cells(f"A{r}:G{r}")
notes = [
    "• Migration script: backend/migrate_utilization.py — uses ALTER TABLE … ADD COLUMN IF NOT EXISTS (idempotent).",
    "• conn.autocommit = True required for DDL in psycopg3 (DDL cannot run inside a transaction block).",
    "• CREATE TABLE IF NOT EXISTS only creates missing tables; does NOT add columns to existing tables — hence ALTER TABLE.",
    "• license_usage existing rows get default values for new columns; update via seed_utilization_data.py.",
    "• po_categorization uses ON CONFLICT (purchasing_document, item) DO UPDATE SET … WHERE tagged_by='SYSTEM' to protect manual overrides.",
]
for i, note in enumerate(notes, r+1):
    cell(ws3, i, 1, note, wrap=True, bg=GREY if i % 2 == 0 else WHITE)
    ws3.merge_cells(f"A{i}:G{i}")
    ws3.row_dimensions[i].height = 22

set_col_widths(ws3, [22, 14, 18, 45, 55, 10, 10])


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Base Utilization KPIs
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Base KPIs")
ws4.sheet_view.showGridLines = False

hdr(ws4, 1, 1, "Base Utilization KPIs — _utilization() function", bg=NAVY, sz=13)
ws4.merge_cells("A1:G1")

hdrs = ["KPI Code", "KPI Name", "Unit", "Source Table(s)", "Logic Summary", "Company Filter", ""]
for c, h in enumerate(hdrs[:6], 1):
    hdr(ws4, 2, c, h, bg=BLUE, sz=10)

base_kpis = [
    ("CAPEX_SPEND_YTD",      "CAPEX Spend YTD",              "INR Cr",  "po_dump",              "SUM(net_order_value) WHERE UPPER(capex_opex_flag)='CAPEX' AND NOT deleted AND doc_date>=FY_START", "YES"),
    ("OPEX_SPEND_YTD",       "OPEX Spend YTD",               "INR Cr",  "po_dump",              "SUM(net_order_value) WHERE UPPER(capex_opex_flag)='OPEX' AND NOT deleted AND doc_date>=FY_START", "YES"),
    ("TOTAL_SPEND_YTD",      "Total PO Spend YTD",           "INR Cr",  "po_dump",              "SUM(net_order_value) NOT deleted AND doc_date>=FY_START", "YES"),
    ("CAPEX_PCT",            "CAPEX % of Total Spend",       "%",       "po_dump",              "CAPEX_SPEND / TOTAL_SPEND × 100", "YES"),
    ("OPEX_PCT",             "OPEX % of Total Spend",        "%",       "po_dump",              "OPEX_SPEND / TOTAL_SPEND × 100", "YES"),
    ("PO_COUNT_YTD",         "PO Count YTD",                 "count",   "po_dump",              "COUNT(*) NOT deleted AND doc_date>=FY_START", "YES"),
    ("AVG_PO_VALUE_YTD",     "Avg PO Value YTD",             "INR Cr",  "po_dump",              "TOTAL_SPEND / PO_COUNT", "YES"),
    ("CAPEX_PO_COUNT",       "CAPEX PO Count YTD",           "count",   "po_dump",              "COUNT where capex_opex_flag='CAPEX'", "YES"),
    ("OPEX_PO_COUNT",        "OPEX PO Count YTD",            "count",   "po_dump",              "COUNT where capex_opex_flag='OPEX'", "YES"),
    ("MTD_SPEND",            "MTD PO Spend",                 "INR Cr",  "po_dump",              "SUM(net_order_value) WHERE doc_date>=MTD_START", "YES"),
    ("PLANT_BREAKDOWN",      "Spend by Plant (JSON)",        "json",    "po_dump",              "GROUP BY plant → top 8 by spend, JSON array [{plant, spend_cr}]", "YES"),
    ("CAT_BREAKDOWN_CAPEX",  "CAPEX Spend by Category (JSON)","json",   "po_dump",              "GROUP BY material_group WHERE CAPEX → JSON [{category, spend_cr}]", "YES"),
    ("CAT_BREAKDOWN_OPEX",   "OPEX Spend by Category (JSON)","json",   "po_dump",              "GROUP BY material_group WHERE OPEX → JSON [{category, spend_cr}]", "YES"),
    ("CAT_BREAKDOWN_ALL",    "All Spend by Category (JSON)", "json",    "po_dump",              "GROUP BY material_group → JSON [{category, spend_cr}]", "YES"),
]

for i, (code, name, unit, src, logic, cc) in enumerate(base_kpis, 3):
    bg = LGREEN if i % 2 == 0 else WHITE
    cell(ws4, i, 1, code, bold=True, bg=bg)
    cell(ws4, i, 2, name, bg=bg)
    cell(ws4, i, 3, unit, bg=bg, halign="center")
    cell(ws4, i, 4, src, bg=bg, halign="center")
    cell(ws4, i, 5, logic, bg=bg, wrap=True)
    cell(ws4, i, 6, cc, bg=bg, halign="center", fg="375623", bold=True)
    ws4.row_dimensions[i].height = 28

border_range(ws4, 2, 2+len(base_kpis), 1, 6)
freeze(ws4, 3, 1)
set_col_widths(ws4, [26, 30, 10, 22, 75, 14, 10])


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Software KPIs
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Software KPIs")
ws5.sheet_view.showGridLines = False

hdr(ws5, 1, 1, "Software License KPIs — _utilization_software() function", bg=NAVY, sz=13)
ws5.merge_cells("A1:G1")

for c, h in enumerate(["KPI Code", "KPI Name", "Unit", "Source Table(s)", "Logic Summary", "Company Filter", ""], 1):
    hdr(ws5, 2, c, h if c <= 6 else "", bg=BLUE, sz=10)

sw_kpis = [
    ("SW_LIC_UTIL_RATE",   "License Utilization Rate",          "%",      "license_usage",            "SUM(active_users)/SUM(total_seats)×100 — measures how much of bought SW is actually used", "replicated"),
    ("SW_UNDERUTIL_COUNT", "Under-utilized Tools Count",        "count",  "license_usage",            "COUNT where active_users/total_seats < 50% — highlights wasteful subscriptions", "replicated"),
    ("SW_TOTAL_LICENSES",  "Total Licensed Seats",              "seats",  "license_usage",            "SUM(total_seats) across all tools", "replicated"),
    ("SW_ACTIVE_USERS",    "Active Users Total",                "users",  "license_usage",            "SUM(active_users)", "replicated"),
    ("SW_UNUSED_SEATS",    "Unused Seats",                      "seats",  "license_usage",            "SUM(total_seats - active_users)", "replicated"),
    ("SW_ANNUAL_COST",     "Annual License Cost",               "INR Cr", "license_usage",            "SUM(annual_cost_inr)/1e7", "replicated"),
    ("SW_COST_PER_USER",   "Cost per Active User",              "INR Lakh","license_usage",           "SW_ANNUAL_COST / SW_ACTIVE_USERS (in Lakh)", "replicated"),
    ("SW_WASTED_COST",     "Wasted License Cost",               "INR Cr", "license_usage",            "SUM(annual_cost_inr × unused_seats/total_seats)/1e7", "replicated"),
    ("SW_RENEWAL_90D",     "Renewals Due in 90 Days",           "count",  "license_usage",            "COUNT where renewal_date BETWEEN today AND today+90", "replicated"),
    ("SW_TOOL_BREAKDOWN",  "Tool-level Usage Detail (JSON)",    "json",   "license_usage",            "Per-tool array: {tool, seats, active, util_pct, cost_cr, renewal_date}", "replicated"),
    ("SW_CAPEX_SPEND",     "Software CAPEX Spend YTD (Cr)",     "INR Cr", "po_dump+po_categorization","SUM(net_order_value) WHERE po_category='SOFTWARE' AND capex_opex_flag='CAPEX' AND NOT deleted", "YES"),
    ("SW_OPEX_SPEND",      "Software OPEX Spend YTD (Cr)",      "INR Cr", "po_dump+po_categorization","SUM(net_order_value) WHERE po_category='SOFTWARE' AND capex_opex_flag='OPEX' AND NOT deleted", "YES"),
    ("SW_VENDOR_CONC",     "Top Vendor Concentration %",        "%",      "po_dump+po_categorization","Top vendor's SW spend / total SW spend × 100", "YES"),
    ("SW_VENDOR_BREAKDOWN","Vendor SW Spend Breakdown (JSON)",  "json",   "po_dump+po_categorization","GROUP BY vendor_name WHERE po_category='SOFTWARE' → top 8, JSON [{vendor, spend_cr, pct}]", "YES"),
]

for i, (code, name, unit, src, logic, cc) in enumerate(sw_kpis, 3):
    bg = LAMBER if i % 2 == 0 else WHITE
    cell(ws5, i, 1, code, bold=True, bg=bg)
    cell(ws5, i, 2, name, bg=bg, wrap=True)
    cell(ws5, i, 3, unit, bg=bg, halign="center")
    cell(ws5, i, 4, src, bg=bg, halign="center", wrap=True)
    cell(ws5, i, 5, logic, bg=bg, wrap=True)
    cc_fg = "7F6000" if cc == "replicated" else "375623"
    cell(ws5, i, 6, cc, bg=bg, halign="center", fg=cc_fg, bold=True)
    ws5.row_dimensions[i].height = 30

border_range(ws5, 2, 2+len(sw_kpis), 1, 6)
freeze(ws5, 3, 1)
set_col_widths(ws5, [26, 34, 12, 26, 80, 14, 10])

# Note row
note_row = 3 + len(sw_kpis)
cell(ws5, note_row, 1, "Note: 'replicated' = same value stored for 1001/1002/1003/ALL (license_usage has no company_code). PO-based SW KPIs (SW_CAPEX_SPEND etc.) DO filter by company.",
     bg=LBLUE, wrap=True, italic=True, fg="2F5496")
ws5.merge_cells(f"A{note_row}:G{note_row}")
ws5.row_dimensions[note_row].height = 25


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — Materials KPIs
# ═══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Materials KPIs")
ws6.sheet_view.showGridLines = False

hdr(ws6, 1, 1, "Materials KPIs — _utilization_materials() function", bg=NAVY, sz=13)
ws6.merge_cells("A1:G1")

for c, h in enumerate(["KPI Code", "KPI Name", "Unit", "Source Table(s)", "Logic Summary", "Company Filter", ""], 1):
    hdr(ws6, 2, c, h if c <= 6 else "", bg=BLUE, sz=10)

mat_kpis = [
    ("MAT_DELIV_UTIL_RATE",        "Material Delivery Utilization %",      "%",      "po_dump+grn_dump+po_categorization",          "SUM(grn_qty)/SUM(po_qty)×100 WHERE movement_type='101' AND debit_credit='S' AND NOT deleted AND category=MATERIAL", "YES"),
    ("MAT_DELIVERY_COMPLETE_PCT",  "Delivery Completed PO %",              "%",      "po_dump",                                      "COUNT(delivery_completed='X') / COUNT(*) × 100 WHERE NOT deleted AND doc_date>=FY", "YES"),
    ("MAT_OPEN_PO_VALUE",          "Open PO Value Not GRN'd (Cr)",         "INR Cr", "pr_po_grn_invoice",                            "SUM(po_net_value - grn_amount) WHERE NOT deleted AND grn_amount < 95% of po_net_value", "YES"),
    ("MAT_CAPEX_SPEND",            "Material CAPEX Spend YTD (Cr)",        "INR Cr", "po_dump+po_categorization",                    "SUM(net_order_value) WHERE po_category='MATERIAL' AND capex_opex_flag='CAPEX' AND NOT deleted AND doc_date>=FY", "YES"),
    ("MAT_OPEX_SPEND",             "Material OPEX Spend YTD (Cr)",         "INR Cr", "po_dump+po_categorization",                    "SUM(net_order_value) WHERE po_category='MATERIAL' AND capex_opex_flag='OPEX' AND NOT deleted AND doc_date>=FY", "YES"),
    ("MAT_LICENSE_COST_TOT",       "Material Licensing Cost Total (Cr)",   "INR Cr", "material_license_cost+po_dump",                "SUM(license_fee_inr) WHERE NOT deleted — total royalty/import licence fees", "YES"),
    ("MAT_LICENSE_BREAKDOWN",      "Material License Cost Breakdown (JSON)","json",  "material_license_cost+po_dump",                "GROUP BY license_type → [{type, cost_cr}]", "YES"),
    ("MAT_3WAY_MATCH",             "3-Way Match Rate (Materials %)",        "%",      "pr_po_grn_invoice+po_categorization",          "COUNT(invoice≈grn within 5%) / COUNT(grn_amount>0) × 100 WHERE po_category=MATERIAL", "YES"),
    ("MAT_VENDOR_FILL_RATE",       "Vendor GRN Fill Rate (JSON)",          "json",   "po_dump+grn_dump",                             "Per-vendor: grn_qty/po_qty×100 → top 8 [{vendor, fill_pct, grn_qty, po_qty}]", "YES"),
    ("MAT_CAT_BREAKDOWN",          "Material Category Spend Breakdown (JSON)","json","po_dump+po_categorization",                    "GROUP BY material_group WHERE po_category=MATERIAL → [{category, spend_cr, pct}]", "YES"),
]

for i, (code, name, unit, src, logic, cc) in enumerate(mat_kpis, 3):
    bg = LGREEN if i % 2 == 0 else WHITE
    cell(ws6, i, 1, code, bold=True, bg=bg)
    cell(ws6, i, 2, name, bg=bg, wrap=True)
    cell(ws6, i, 3, unit, bg=bg, halign="center")
    cell(ws6, i, 4, src, bg=bg, halign="center", wrap=True)
    cell(ws6, i, 5, logic, bg=bg, wrap=True)
    cell(ws6, i, 6, cc, bg=bg, halign="center", fg="375623", bold=True)
    ws6.row_dimensions[i].height = 32

border_range(ws6, 2, 2+len(mat_kpis), 1, 6)
freeze(ws6, 3, 1)
set_col_widths(ws6, [28, 35, 12, 32, 80, 14, 10])


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — Auto-Categorization
# ═══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Auto-Categorization")
ws7.sheet_view.showGridLines = False

hdr(ws7, 1, 1, "Auto-Categorization Engine — _auto_categorize()", bg=NAVY, sz=13)
ws7.merge_cells("A1:E1")

r = 2
cell(ws7, r, 1, "Purpose", bold=True, bg=LBLUE)
cell(ws7, r, 2, "Scans material_description in po_dump using keyword regex patterns and inserts SYSTEM tags into po_categorization. Manual USER tags are never overwritten.", bg=LBLUE, wrap=True)
ws7.merge_cells(f"B{r}:E{r}")
ws7.row_dimensions[r].height = 25

r += 1
cell(ws7, r, 1, "Conflict rule", bold=True, bg=LBLUE)
cell(ws7, r, 2, "ON CONFLICT (purchasing_document, item) DO UPDATE SET … WHERE tagged_by='SYSTEM'  — so user overrides survive recompute", bg=LBLUE, wrap=True)
ws7.merge_cells(f"B{r}:E{r}")
ws7.row_dimensions[r].height = 22

r += 1
cell(ws7, r, 1, "When called", bold=True, bg=LBLUE)
cell(ws7, r, 2, "Only when company_code='ALL' inside _utilization() — prevents N redundant runs per compute cycle", bg=LBLUE, wrap=True)
ws7.merge_cells(f"B{r}:E{r}")
ws7.row_dimensions[r].height = 22

r += 2
hdr(ws7, r, 1, "SOFTWARE Keyword Patterns", bg=BLUE)
ws7.merge_cells(f"A{r}:E{r}")
r += 1
for c, h in enumerate(["Pattern (regex, case-insensitive)", "Maps to category", "", "", ""], 1):
    if h:
        hdr(ws7, r, c, h, bg=LBLUE, fg="000000", sz=10)
r += 1

sw_patterns = [
    (r"licen[sc]", "SOFTWARE"), (r"software|saas|cloud subscri", "SOFTWARE"),
    (r"erp|sap|oracle|salesforce|microsoft\s+365|office\s+365", "SOFTWARE"),
    (r"annual\s+maintenance.*contract|amc", "SOFTWARE"),
    (r"it\s+service|managed\s+service|helpdesk", "SOFTWARE"),
]
for pat, cat in sw_patterns:
    bg = LAMBER if sw_patterns.index((pat, cat)) % 2 == 0 else WHITE
    cell(ws7, r, 1, pat, italic=True, bg=bg)
    cell(ws7, r, 2, cat, bold=True, bg=bg, fg=AMBER)
    ws7.merge_cells(f"B{r}:E{r}")
    r += 1

r += 1
hdr(ws7, r, 1, "MATERIAL Keyword Patterns (overrides SERVICE default)", bg=BLUE)
ws7.merge_cells(f"A{r}:E{r}")
r += 1
mat_patterns = [
    (r"raw\s+material|consumable|spare\s+part|hardware\s+component", "MATERIAL"),
    (r"steel|copper|alumin|chemical|polymer|resin|rubber|plastic|paper", "MATERIAL"),
    (r"equipment\s+purchase|machinery|tool\s+&\s+die", "MATERIAL"),
]
for pat, cat in mat_patterns:
    bg = LGREEN if mat_patterns.index((pat, cat)) % 2 == 0 else WHITE
    cell(ws7, r, 1, pat, italic=True, bg=bg)
    cell(ws7, r, 2, cat, bold=True, bg=bg, fg=GREEN)
    ws7.merge_cells(f"B{r}:E{r}")
    r += 1

r += 1
hdr(ws7, r, 1, "Fallback / Default", bg=BLUE)
ws7.merge_cells(f"A{r}:E{r}")
r += 1
cell(ws7, r, 1, "No keyword match", bg=GREY)
cell(ws7, r, 2, "SERVICE (default category for unrecognized material descriptions)", bg=GREY, wrap=True)
ws7.merge_cells(f"B{r}:E{r}")

set_col_widths(ws7, [50, 30, 12, 12, 12])


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 8 — Data Flow
# ═══════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Data Flow")
ws8.sheet_view.showGridLines = False

hdr(ws8, 1, 1, "End-to-End Data Flow — Utilization Dashboard", bg=NAVY, sz=13)
ws8.merge_cells("A1:C1")

steps = [
    ("1", "CSV Upload",          "User uploads PO/GRN/Invoice CSVs via /upload page → load_data.py ingests into PostgreSQL tables (po_dump, grn_dump, etc.)"),
    ("2", "Manual Seeding",      "seed_utilization_data.py populates: license_usage (10 SW tools), profit_center_master (8 PCs), pc_budget (12 rows), material_license_cost (32 rows)"),
    ("3", "Auto-Categorize",     "_auto_categorize() — keyword regex scan on po_dump.material_description → inserts SYSTEM tags into po_categorization (USER tags protected)"),
    ("4", "Company Loop",        "compute_all() runs:\n  for cc in [1001, 1002, 1003]: _utilization(cc)\n  _utilization('ALL')\nEach call: _utilization_software(cc) + _utilization_materials(cc)"),
    ("5", "KPI Upsert",          "All results → _upsert(conn, 'utilization', kpi_code, …, company_code=cc) → kpi_results table with ON CONFLICT DO UPDATE"),
    ("6", "API Serve",           "FastAPI GET /api/kpi/utilization?company_code=1001 → returns JSON array of {kpi_code, value_numeric, value_text, unit, trend}"),
    ("7", "Company List API",    "GET /api/kpi/utilization/companies → returns {companies: ['1001','1002','1003','ALL']} → CompanyFilter renders in UI"),
    ("8", "Frontend Fetch",      "useKpiCompanies('utilization') → company select\nuseKpi('utilization', company) → all KPIs for chosen company\nuseKpiValue('utilization', code, company) → single KPI value"),
    ("9", "UI Rendering",        "UtilizationDashboard renders:\n  → SpendRow / CountRow / PlantBreakdown / CategoryBreakdown (Base KPIs)\n  → SoftwareSection (SW_* codes)\n  → MaterialsSection (MAT_* codes)"),
    ("10","ⓘ Info Hover",        "Each KPI card has KpiInfoHover → reads KPI_META[kpi_code] from kpiMeta.ts → shows description + formula + collapsible SQL on hover"),
]

for c, h in enumerate(["Step", "Stage", "Detail"], 1):
    hdr(ws8, 2, c, h, bg=BLUE, sz=11)

for i, (step, stage, detail) in enumerate(steps, 3):
    bg = GREY if i % 2 == 0 else WHITE
    cell(ws8, i, 1, step, bg=bg, bold=True, halign="center")
    cell(ws8, i, 2, stage, bg=bg, bold=True)
    cell(ws8, i, 3, detail, bg=bg, wrap=True)
    ws8.row_dimensions[i].height = 45

border_range(ws8, 2, 2+len(steps), 1, 3)
set_col_widths(ws8, [8, 22, 90])

# save
out = "Utilization_Schema_and_KPIs.xlsx"
wb.save(out)
print(f"Saved: {out}")
