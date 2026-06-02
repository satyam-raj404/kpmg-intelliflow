"""
IntelliSource P2P — KPI Documentation Export
Generates IntelliSource_KPI_Documentation.xlsx with all implemented KPIs,
formulas, source tables, targets, and anomaly rules.
"""
import sys
import os
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent / "backend"))

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ── Color Palette ─────────────────────────────────────────────────────────────
KPMG_BLUE       = "00338D"
KPMG_BLUE_LIGHT = "D6E0F5"
KPMG_CYAN       = "009FDA"
KPMG_CYAN_LIGHT = "D6F0FA"
KPMG_TEAL       = "00A3A1"
KPMG_TEAL_LIGHT = "D6F0EF"
KPMG_PURPLE     = "470A68"
KPMG_PURP_LIGHT = "EAD9F0"
WHITE           = "FFFFFF"
LIGHT_GREY      = "F5F5F5"
MID_GREY        = "D9D9D9"
DARK_GREY       = "595959"
GREEN           = "009A44"
GREEN_LIGHT     = "D6F0E1"
AMBER           = "F5A623"
AMBER_LIGHT     = "FDF0D9"
RED             = "D0021B"
RED_LIGHT       = "FAD6DA"

DASHBOARD_COLORS = {
    "Procurement":  (KPMG_BLUE,   KPMG_BLUE_LIGHT,  "PROCUREMENT DASHBOARD"),
    "Financial":    (KPMG_TEAL,   KPMG_TEAL_LIGHT,  "FINANCIAL DASHBOARD"),
    "Leadership":   (KPMG_PURPLE, KPMG_PURP_LIGHT,  "LEADERSHIP DASHBOARD"),
    "Vendor":       (KPMG_CYAN,   KPMG_CYAN_LIGHT,  "VENDOR PERFORMANCE DASHBOARD"),
    "Utilization":  ("1D6F42",    "D6EDDE",          "UTILIZATION DASHBOARD"),
}

def fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def font(bold=False, color=WHITE, size=10, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)

def border(style="thin"):
    s = Side(style=style, color=MID_GREY)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def style_header_row(ws, row, bg_hex, fg_hex=WHITE, font_size=10, bold=True):
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill     = fill(bg_hex)
            cell.font     = font(bold=bold, color=fg_hex, size=font_size)
            cell.border   = border()
            cell.alignment = align("center", "center", wrap=True)

def style_data_row(ws, row_num, alt=False, cols=None):
    bg = LIGHT_GREY if alt else WHITE
    for cell in ws[row_num]:
        if cols and cell.column not in cols:
            continue
        cell.fill      = fill(bg)
        cell.font      = Font(color="000000", size=9, name="Calibri")
        cell.border    = border()
        cell.alignment = align("left", "center", wrap=True)

# ── KPI DATA ──────────────────────────────────────────────────────────────────

PROCUREMENT_KPIS = [
    {
        "kpi_code":       "TOTAL_PO_VALUE_MTD",
        "kpi_name":       "Total PO Value (MTD)",
        "business_q":     "What is the total spend committed in POs this month?",
        "formula_plain":  "SUM(net_order_value) WHERE document_date >= month_start AND deletion_indicator NOT IN ('L','X')",
        "formula_sql":    "SELECT SUM(CAST(net_order_value AS REAL))\nFROM po_dump\nWHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))\n  AND document_date >= MTD_START",
        "source_tables":  "po_dump",
        "source_fields":  "net_order_value, document_date, deletion_indicator",
        "unit":           "INR",
        "target":         "—",
        "rag":            "No threshold",
        "notes":          "MTD_START = first day of current month based on latest document_date in po_dump",
    },
    {
        "kpi_code":       "ACTIVE_PO_COUNT",
        "kpi_name":       "Active PO Count (MTD)",
        "business_q":     "How many POs are currently open (not complete, not deleted)?",
        "formula_plain":  "COUNT(DISTINCT purchasing_document + item) WHERE delivery_completed='' AND deletion_indicator NOT IN ('L','X') AND document_date >= month_start",
        "formula_sql":    "SELECT COUNT(DISTINCT purchasing_document || '|' || item)\nFROM po_dump\nWHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))\n  AND (delivery_completed IS NULL OR delivery_completed = '')\n  AND document_date >= MTD_START",
        "source_tables":  "po_dump",
        "source_fields":  "purchasing_document, item, deletion_indicator, delivery_completed",
        "unit":           "Count",
        "target":         "—",
        "rag":            "No threshold",
        "notes":          "Counts distinct PO line keys. delivery_completed='X' means fully delivered.",
    },
    {
        "kpi_code":       "HIGH_VALUE_PO_COUNT",
        "kpi_name":       "High-Value PO Count",
        "business_q":     "How many POs exceed the configurable high-value threshold?",
        "formula_plain":  "COUNT(DISTINCT purchasing_document) WHERE net_order_value > HIGH_VALUE_PO_THRESHOLD AND deletion_indicator NOT IN ('L','X')",
        "formula_sql":    "SELECT COUNT(DISTINCT purchasing_document)\nFROM po_dump\nWHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))\n  AND CAST(net_order_value AS REAL) > [HIGH_VALUE_PO_THRESHOLD from kpi_config]",
        "source_tables":  "po_dump, kpi_config",
        "source_fields":  "net_order_value, deletion_indicator",
        "unit":           "Count",
        "target":         "Configurable (user sets in Leadership dashboard header)",
        "rag":            "User-defined",
        "notes":          "Threshold stored in kpi_config.HIGH_VALUE_PO_THRESHOLD. Default = ₹1 Cr. PUT /api/kpi-config/HIGH_VALUE_PO_THRESHOLD to change.",
    },
    {
        "kpi_code":       "PR_TO_PO_DAYS",
        "kpi_name":       "Avg PR-to-PO Conversion Time",
        "business_q":     "How many days on average from PR approval to PO creation at LINE level?",
        "formula_plain":  "AVG(DATEDIFF(po.document_date, pr.release_date)) at item level — joined via purchase_requisition + item_of_requisition",
        "formula_sql":    "SELECT AVG(CAST(pr_to_po_days AS REAL))\nFROM pr_po_grn_invoice\nWHERE pr_to_po_days IS NOT NULL\n  AND pr_to_po_days >= 0\n  AND purchase_requisition IS NOT NULL\n  AND purchasing_document IS NOT NULL",
        "source_tables":  "pr_po_grn_invoice (computed from pr_dump + po_dump)",
        "source_fields":  "pr_dump.release_date, po_dump.document_date joined on purchase_requisition+item_of_requisition",
        "unit":           "Days",
        "target":         "≤ 5 days",
        "rag":            "GREEN ≤5, AMBER 5-10, RED >10",
        "notes":          "ITEM-LEVEL join: pr.purchase_requisition = po.purchase_requisition AND pr.item_of_requisition = po.item_of_requisition. Takes first PO date per PR line.",
    },
    {
        "kpi_code":       "PO_APPROVAL_CYCLE",
        "kpi_name":       "PO Approval Cycle Time",
        "business_q":     "How many days from PO creation to PO release/approval?",
        "formula_plain":  "AVG(DATEDIFF(change_log[FRGZU/FRGKE=X].change_date, po.document_date))",
        "formula_sql":    "SELECT AVG(CAST(julianday(cl.change_date) - julianday(po.document_date) AS REAL))\nFROM po_dump po\nJOIN change_log cl ON cl.object_id = po.purchasing_document\n  AND cl.object_class = 'EINKBELEG'\n  AND cl.field_name IN ('FRGZU','FRGKE')\n  AND cl.change_indicator = 'U'\n  AND cl.new_value = 'X'\nWHERE po.release_indicator = 'X'",
        "source_tables":  "po_dump, change_log",
        "source_fields":  "po_dump.document_date, change_log.change_date WHERE field_name='FRGZU' AND change_indicator='U' AND new_value='X'",
        "unit":           "Days",
        "target":         "≤ 3 days",
        "rag":            "GREEN ≤3, AMBER 3-7, RED >7",
        "notes":          "Release date sourced from change_log (FRGZU/FRGKE field, U-type change, new_value='X'). SAP transaction ME29N.",
    },
    {
        "kpi_code":       "PO_DELETION_MTD",
        "kpi_name":       "PO Deletion Frequency (MTD)",
        "business_q":     "How many POs were deleted/cancelled this month?",
        "formula_plain":  "COUNT(DISTINCT purchasing_document) WHERE deletion_indicator='L' AND document_date >= month_start",
        "formula_sql":    "SELECT COUNT(DISTINCT purchasing_document)\nFROM po_dump\nWHERE deletion_indicator = 'L'\n  AND document_date >= MTD_START",
        "source_tables":  "po_dump",
        "source_fields":  "deletion_indicator (L=deleted), document_date",
        "unit":           "Count",
        "target":         "≤ 5 per month",
        "rag":            "GREEN ≤5, RED >5",
        "notes":          "SAP field EKPO-LOEKZ. L = line item deleted.",
    },
    {
        "kpi_code":       "PO_AMENDMENT_RATE",
        "kpi_name":       "PO Amendment Rate",
        "business_q":     "What % of POs were modified after creation (excluding initial inserts)?",
        "formula_plain":  "COUNT(DISTINCT PO with change_indicator='U' in change_log) / COUNT(DISTINCT all POs) × 100",
        "formula_sql":    "SELECT COUNT(DISTINCT cl.object_id) * 100.0\n       / NULLIF(COUNT(DISTINCT po.purchasing_document), 0)\nFROM po_dump po\nLEFT JOIN change_log cl\n  ON cl.object_id = po.purchasing_document\n AND cl.object_class = 'EINKBELEG'\n AND cl.change_indicator = 'U'          -- U=Update only\n AND cl.field_name NOT IN ('FRGZU','FRGKE')  -- exclude release approvals\nWHERE po.deletion_indicator IS NULL OR po.deletion_indicator = ''",
        "source_tables":  "po_dump, change_log",
        "source_fields":  "change_log.change_indicator='U' (Updates only — excludes I=Insert, D=Delete)",
        "unit":           "%",
        "target":         "< 15%",
        "rag":            "GREEN <15%, AMBER 15-25%, RED >25%",
        "notes":          "CRITICAL: filters change_indicator='U' only. change_indicator='I' = initial creation (excluded). Releases (FRGZU/FRGKE) also excluded from amendment count.",
    },
    {
        "kpi_code":       "OPEN_PR_AGING",
        "kpi_name":       "Open PR Aging > 7 Days",
        "business_q":     "How many approved PR lines have no PO after 7+ days?",
        "formula_plain":  "COUNT(PR lines WHERE release_status='X' AND no matching PO at item level AND days_since_release > 7)",
        "formula_sql":    "SELECT COUNT(DISTINCT pr.purchase_requisition || '|' || pr.item_of_requisition)\nFROM pr_dump pr\nWHERE pr.release_status IN ('X','XX','XXX','XXXX','XXXXX')\n  AND (pr.deletion_indicator IS NULL OR pr.deletion_indicator = '')\n  AND pr.release_date IS NOT NULL\n  AND julianday('now') - julianday(pr.release_date) > 7\n  AND NOT EXISTS (\n    SELECT 1 FROM po_dump po\n    WHERE po.purchase_requisition = pr.purchase_requisition\n      AND po.item_of_requisition  = pr.item_of_requisition\n  )",
        "source_tables":  "pr_dump, po_dump",
        "source_fields":  "pr_dump.release_status, release_date; po_dump.purchase_requisition, item_of_requisition",
        "unit":           "Count",
        "target":         "≤ 10 PR lines",
        "rag":            "GREEN ≤10, AMBER 10-20, RED >20",
        "notes":          "Item-level check: PR line has no matching PO line via purchase_requisition + item_of_requisition composite key.",
    },
    {
        "kpi_code":       "TOTAL_PO_VALUE_YTD",
        "kpi_name":       "Total PO Value (YTD)",
        "business_q":     "What is total committed PO spend for current Indian FY?",
        "formula_plain":  "SUM(net_order_value) WHERE document_date >= FY_start (April 1) AND deletion_indicator NOT IN ('L','X')",
        "formula_sql":    "SELECT SUM(CAST(net_order_value AS REAL))\nFROM po_dump\nWHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))\n  AND document_date >= '{FY_YEAR}-04-01'",
        "source_tables":  "po_dump",
        "source_fields":  "net_order_value, document_date, deletion_indicator",
        "unit":           "INR",
        "target":         "Within annual budget",
        "rag":            "No threshold",
        "notes":          "Indian FY = April 1 to March 31. FY start calculated from latest data date.",
    },
    {
        "kpi_code":       "PO_LINE_COUNT_YTD",
        "kpi_name":       "PO Line Count (YTD)",
        "business_q":     "How many PO lines have been created this FY?",
        "formula_plain":  "COUNT(*) FROM po_dump WHERE document_date >= FY_start AND not deleted",
        "formula_sql":    "SELECT COUNT(*) FROM po_dump\nWHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))\n  AND document_date >= '{FY_YEAR}-04-01'",
        "source_tables":  "po_dump",
        "source_fields":  "document_date, deletion_indicator",
        "unit":           "Count",
        "target":         "—",
        "rag":            "Track trend",
        "notes":          "Counts PO lines (not distinct PO headers).",
    },
]

FINANCIAL_KPIS = [
    {
        "kpi_code":       "TOTAL_SPEND_YTD",
        "kpi_name":       "Total Spend YTD (Invoices)",
        "business_q":     "What is total actual spend invoiced this FY (PO + non-PO invoices)?",
        "formula_plain":  "SUM(amount_local_ccy) FROM invoice_dump WHERE document_type IN ('RE','KR') AND amount > 0 AND posting_date >= FY_start",
        "formula_sql":    "SELECT SUM(\n  CASE WHEN document_type IN ('RE','KR')\n       THEN CAST(amount_local_ccy AS REAL)\n       ELSE 0 END\n)\nFROM invoice_dump\nWHERE document_type IN ('RE','KR')\n  AND CAST(amount_local_ccy AS REAL) > 0\n  AND posting_date >= FY_START",
        "source_tables":  "invoice_dump",
        "source_fields":  "document_type (RE=PO invoice, KR=non-PO invoice), amount_local_ccy, posting_date",
        "unit":           "INR",
        "target":         "Within approved budget",
        "rag":            "Compared against BUDGET_ALLOCATION",
        "notes":          "RE = invoice receipt (PO-linked from BSAK). KR = vendor invoice (non-PO, e.g. rent, utilities). RN = cancellation (excluded). D/C adjustment: positive amounts only (S indicator).",
    },
    {
        "kpi_code":       "INVOICE_CANCELLATION_RATE",
        "kpi_name":       "Invoice Cancellation Rate (%)",
        "business_q":     "What % of invoices are being cancelled/reversed?",
        "formula_plain":  "COUNT(invoices WHERE document_type='RN' OR amount<0) / COUNT(all invoices) × 100",
        "formula_sql":    "SELECT\n  COUNT(CASE WHEN document_type='RN' OR CAST(amount_local_ccy AS REAL)<0 THEN 1 END)\n  * 100.0 / NULLIF(COUNT(*), 0)\nFROM invoice_dump",
        "source_tables":  "invoice_dump",
        "source_fields":  "document_type (RN=cancellation), amount_local_ccy",
        "unit":           "%",
        "target":         "< 5%",
        "rag":            "GREEN <5%, AMBER 5-10%, RED >10%",
        "notes":          "RN = invoice reversal document type in SAP. Negative amount_local_ccy also treated as cancellation.",
    },
    {
        "kpi_code":       "THREE_WAY_MATCH_RATE",
        "kpi_name":       "3-Way Match Success Rate (%)",
        "business_q":     "What % of PO invoices match GRN quantity within 5% tolerance?",
        "formula_plain":  "COUNT(lines WHERE |net_grn_qty - invoice_qty| / invoice_qty ≤ 5%) / COUNT(invoiced lines) × 100",
        "formula_sql":    "SELECT\n  COUNT(CASE\n    WHEN ABS(COALESCE(f.grn_quantity,0) - COALESCE(f.invoice_quantity,0))\n         / NULLIF(ABS(COALESCE(f.invoice_quantity,0)), 0) <= 0.05\n    THEN 1 END) * 100.0\n  / NULLIF(COUNT(CASE WHEN f.invoice_quantity IS NOT NULL THEN 1 END), 0)\nFROM pr_po_grn_invoice f\nWHERE f.invoice_quantity IS NOT NULL AND f.grn_quantity IS NOT NULL",
        "source_tables":  "pr_po_grn_invoice (computed from po_dump + grn_dump + po_invoice_dump)",
        "source_fields":  "grn_quantity (net: S - H), invoice_quantity (net: S - H)",
        "unit":           "%",
        "target":         "> 95%",
        "rag":            "GREEN >95%, AMBER 85-95%, RED <85%",
        "notes":          "GRN quantity = SUM(qty WHERE debit_credit='S') - SUM(qty WHERE debit_credit='H'). Net of returns. Tolerance: ±5%.",
    },
    {
        "kpi_code":       "INVOICE_PROCESSING_DAYS",
        "kpi_name":       "Invoice Processing Cycle Time",
        "business_q":     "How many days from vendor invoice date to payment clearing?",
        "formula_plain":  "AVG(DATEDIFF(payment.clearing_date, invoice.vendor_invoice_date))",
        "formula_sql":    "SELECT AVG(julianday(p.clearing_date) - julianday(i.vendor_invoice_date))\nFROM invoice_dump i\nJOIN payment_dump p ON i.clearing_doc = p.payment_doc\nWHERE i.vendor_invoice_date IS NOT NULL\n  AND p.clearing_date IS NOT NULL",
        "source_tables":  "invoice_dump, payment_dump",
        "source_fields":  "invoice_dump.vendor_invoice_date (BLDAT), payment_dump.clearing_date (AUGDT), invoice_dump.clearing_doc = payment_dump.payment_doc",
        "unit":           "Days",
        "target":         "≤ 5 days",
        "rag":            "GREEN ≤5, AMBER 5-15, RED >15",
        "notes":          "Clearing link: invoice_dump.clearing_doc = payment_dump.payment_doc (SAP AUGBL field).",
    },
    {
        "kpi_code":       "ON_TIME_PAYMENT_RATE",
        "kpi_name":       "Payment On-Time Rate (%)",
        "business_q":     "What % of invoices were paid on or before due date?",
        "formula_plain":  "COUNT(payments WHERE payment.posting_date ≤ invoice.due_date) / COUNT(all payments) × 100",
        "formula_sql":    "SELECT\n  COUNT(CASE WHEN p.posting_date <= i.due_date THEN 1 END) * 100.0\n  / NULLIF(COUNT(*), 0)\nFROM payment_dump p\nJOIN invoice_dump i ON p.cleared_invoice = i.invoice_doc\nWHERE i.due_date IS NOT NULL",
        "source_tables":  "payment_dump, invoice_dump",
        "source_fields":  "payment_dump.posting_date, invoice_dump.due_date",
        "unit":           "%",
        "target":         "> 90%",
        "rag":            "GREEN >90%, AMBER 75-90%, RED <75%",
        "notes":          "due_date = invoice_dump.baseline_date (ZFBDT) + invoice_dump.days_1 (ZBD3T). Join: payment_dump.cleared_invoice = invoice_dump.invoice_doc.",
    },
    {
        "kpi_code":       "DPO",
        "kpi_name":       "Days Payable Outstanding (DPO)",
        "business_q":     "On average how many days are we taking to pay vendors (from invoice posting)?",
        "formula_plain":  "AVG(DATEDIFF(payment.posting_date, invoice.posting_date))",
        "formula_sql":    "SELECT AVG(julianday(p.posting_date) - julianday(i.posting_date))\nFROM payment_dump p\nJOIN invoice_dump i ON p.cleared_invoice = i.invoice_doc\nWHERE i.posting_date IS NOT NULL",
        "source_tables":  "payment_dump, invoice_dump",
        "source_fields":  "payment_dump.posting_date (BUDAT), invoice_dump.posting_date (BUDAT)",
        "unit":           "Days",
        "target":         "Match payment terms (N030=30d, N045=45d, N060=60d)",
        "rag":            "Compared to payment_terms days",
        "notes":          "Positive = paid later than invoice date. Different from EARLY_PAYMENT (vs due_date). DPO uses invoice_posting_date as baseline.",
    },
    {
        "kpi_code":       "OPEN_INVOICE_VALUE",
        "kpi_name":       "Open Invoice Value",
        "business_q":     "How much do we owe vendors in unpaid invoices?",
        "formula_plain":  "SUM(amount_local_ccy) FROM invoice_dump WHERE clearing_doc IS NULL AND amount > 0",
        "formula_sql":    "SELECT SUM(CAST(amount_local_ccy AS REAL))\nFROM invoice_dump\nWHERE (clearing_doc IS NULL OR clearing_doc = '')\n  AND CAST(amount_local_ccy AS REAL) > 0",
        "source_tables":  "invoice_dump",
        "source_fields":  "amount_local_ccy, clearing_doc (AUGBL — NULL if unpaid)",
        "unit":           "INR",
        "target":         "Minimize / < ₹5 Cr in 90+ days bucket",
        "rag":            "Monitor aging buckets",
        "notes":          "Unpaid = clearing_doc IS NULL or empty. Also tracked in OPEN_INVOICE_AGING_BUCKETS (0-30/31-60/61-90/90+ days).",
    },
    {
        "kpi_code":       "EARLY_PAYMENT_COUNT",
        "kpi_name":       "Early Payment Count",
        "business_q":     "How many payments were made before the due date?",
        "formula_plain":  "COUNT(payments WHERE payment.posting_date < invoice.due_date)",
        "formula_sql":    "SELECT COUNT(*)\nFROM payment_dump p\nJOIN invoice_dump i ON p.cleared_invoice = i.invoice_doc\nWHERE p.posting_date < i.due_date",
        "source_tables":  "payment_dump, invoice_dump",
        "source_fields":  "payment_dump.posting_date, invoice_dump.due_date",
        "unit":           "Count",
        "target":         "Maximize (early = discount opportunities)",
        "rag":            "Higher is better",
        "notes":          "Pair with LATE_PAYMENT_COUNT for full picture. due_date = baseline_date + days_1.",
    },
    {
        "kpi_code":       "LATE_PAYMENT_COUNT",
        "kpi_name":       "Late Payment Count",
        "business_q":     "How many invoices were paid after the due date?",
        "formula_plain":  "COUNT(payments WHERE payment.posting_date > invoice.due_date)",
        "formula_sql":    "SELECT COUNT(*)\nFROM payment_dump p\nJOIN invoice_dump i ON p.cleared_invoice = i.invoice_doc\nWHERE p.posting_date > i.due_date",
        "source_tables":  "payment_dump, invoice_dump",
        "source_fields":  "payment_dump.posting_date, invoice_dump.due_date",
        "unit":           "Count",
        "target":         "< 10% of total payments",
        "rag":            "GREEN <10%, RED >20%",
        "notes":          "Late payments risk vendor relationship and may attract penalty interest.",
    },
    {
        "kpi_code":       "TOTAL_PAYMENTS_YTD",
        "kpi_name":       "Total Payments (YTD)",
        "business_q":     "How much cash has actually left the company this FY?",
        "formula_plain":  "SUM(amount_local_ccy) FROM payment_dump WHERE posting_date >= FY_start",
        "formula_sql":    "SELECT SUM(CAST(amount_local_ccy AS REAL))\nFROM payment_dump\nWHERE posting_date >= '{FY_YEAR}-04-01'",
        "source_tables":  "payment_dump",
        "source_fields":  "amount_local_ccy, posting_date",
        "unit":           "INR",
        "target":         "—",
        "rag":            "Informational",
        "notes":          "Actual cash outflow. Different from TOTAL_SPEND_YTD which is invoice-based (committed spend).",
    },
    {
        "kpi_code":       "OPEN_INVOICE_AGING_BUCKETS",
        "kpi_name":       "Open Invoice Aging Buckets",
        "business_q":     "How is outstanding payables distributed across age buckets?",
        "formula_plain":  "SUM(amount) bucketed by days since posting_date: 0-30 / 31-60 / 61-90 / 90+",
        "formula_sql":    "SELECT\n  SUM(CASE WHEN julianday('now') - julianday(posting_date) <= 30 THEN amount ELSE 0 END),\n  SUM(CASE WHEN ...31-60... THEN amount ELSE 0 END),\n  SUM(CASE WHEN ...61-90... THEN amount ELSE 0 END),\n  SUM(CASE WHEN julianday('now') - julianday(posting_date) > 90 THEN amount ELSE 0 END)\nFROM invoice_dump\nWHERE clearing_doc IS NULL AND amount > 0",
        "source_tables":  "invoice_dump",
        "source_fields":  "amount_local_ccy, posting_date, clearing_doc",
        "unit":           "JSON (INR per bucket)",
        "target":         "90+ bucket < ₹5 Cr",
        "rag":            "90+ bucket: GREEN <5Cr, RED >10Cr",
        "notes":          "Returned as JSON: {\"0-30d\": X, \"31-60d\": Y, \"61-90d\": Z, \"90+d\": W}",
    },
]

LEADERSHIP_KPIS = [
    {
        "kpi_code":       "TOTAL_SPEND_YTD",
        "kpi_name":       "Total Procurement Value (YTD)",
        "business_q":     "What is total committed procurement spend (PO value) this FY?",
        "formula_plain":  "SUM(net_order_value) FROM po_dump WHERE document_date >= FY_start AND not deleted",
        "formula_sql":    "SELECT SUM(CAST(net_order_value AS REAL))\nFROM po_dump\nWHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))\n  AND document_date >= '{FY_YEAR}-04-01'",
        "source_tables":  "po_dump",
        "source_fields":  "net_order_value, document_date, deletion_indicator",
        "unit":           "INR",
        "target":         "Within annual plan",
        "rag":            "vs Budget",
        "notes":          "Committed spend (PO value). Actual cash out = TOTAL_PAYMENTS_YTD in Financial dashboard.",
    },
    {
        "kpi_code":       "MAVERICK_BUY_RATE",
        "kpi_name":       "Maverick PO Rate (%)",
        "business_q":     "What % of POs were created without an upstream PR (off-process buying)?",
        "formula_plain":  "COUNT(POs WHERE purchase_requisition IS NULL) / COUNT(all POs) × 100",
        "formula_sql":    "SELECT\n  COUNT(CASE WHEN is_maverick = 1 THEN 1 END) * 100.0\n  / NULLIF(COUNT(*), 0)\nFROM pr_po_grn_invoice\nWHERE purchasing_document IS NOT NULL",
        "source_tables":  "pr_po_grn_invoice (derived from po_dump)",
        "source_fields":  "is_maverick flag (1 = po_dump.purchase_requisition IS NULL or '')",
        "unit":           "%",
        "target":         "< 5%",
        "rag":            "GREEN <5%, AMBER 5-15%, RED >15%",
        "notes":          "Also surfaced as MAVERICK_BUY anomaly in P2P lifecycle. SAP field EKPO-BANFN NULL = no PR.",
    },
    {
        "kpi_code":       "E2E_CYCLE_TIME",
        "kpi_name":       "End-to-End P2P Cycle Time",
        "business_q":     "How long from PR release to payment on average?",
        "formula_plain":  "AVG(total_cycle_days) — PR release_date to first payment_date",
        "formula_sql":    "SELECT AVG(CAST(total_cycle_days AS REAL))\nFROM pr_po_grn_invoice\nWHERE total_cycle_days IS NOT NULL AND total_cycle_days > 0",
        "source_tables":  "pr_po_grn_invoice",
        "source_fields":  "pr_release_date, payment via po_invoice_dump → invoice_dump → payment_dump chain",
        "unit":           "Days",
        "target":         "≤ 45 days",
        "rag":            "GREEN ≤45, AMBER 45-60, RED >60",
        "notes":          "total_cycle_days = MIN(payment_date) - PR release_date. For maverick POs uses po_document_date as start.",
    },
    {
        "kpi_code":       "VENDOR_CONCENTRATION",
        "kpi_name":       "Top-3 Vendor Spend Concentration (%)",
        "business_q":     "What % of total spend goes to the top 3 vendors?",
        "formula_plain":  "SUM(top-3 vendors' net_order_value) / SUM(all vendors' net_order_value) × 100",
        "formula_sql":    "SELECT SUM(sub.spend) * 100.0 / NULLIF(total_spend, 0)\nFROM (\n  SELECT SUM(CAST(net_order_value AS REAL)) AS spend\n  FROM po_dump WHERE deletion_indicator NOT IN ('L','X')\n  GROUP BY vendor ORDER BY spend DESC LIMIT 3\n) sub",
        "source_tables":  "po_dump",
        "source_fields":  "vendor, net_order_value, deletion_indicator",
        "unit":           "%",
        "target":         "< 40%",
        "rag":            "GREEN <40%, AMBER 40-60%, RED >60%",
        "notes":          "High concentration = single-vendor dependency risk. Also track material-group-level single-source procurement.",
    },
    {
        "kpi_code":       "NEGOTIATION_SAVINGS",
        "kpi_name":       "Negotiation Savings YTD",
        "business_q":     "How much have we saved vs PR estimated price through negotiation?",
        "formula_plain":  "SUM((pr.valuation_price - po.net_order_price) × po.order_quantity) WHERE savings > 0",
        "formula_sql":    "SELECT SUM(\n  (CAST(f.pr_value AS REAL) - CAST(f.po_net_price AS REAL))\n  * CAST(f.po_quantity AS REAL)\n)\nFROM pr_po_grn_invoice f\nWHERE f.pr_value IS NOT NULL\n  AND f.po_net_price IS NOT NULL\n  AND (CAST(f.pr_value AS REAL) - CAST(f.po_net_price AS REAL)) > 0",
        "source_tables":  "pr_po_grn_invoice",
        "source_fields":  "pr_value (valuation_price per unit), po_net_price, po_quantity",
        "unit":           "INR",
        "target":         "Maximize",
        "rag":            "Higher is better",
        "notes":          "Only counts lines where PO price < PR estimated price (savings). Lines where PO > PR are price deviations (anomaly).",
    },
    {
        "kpi_code":       "SUPPLY_RISK_SCORE",
        "kpi_name":       "Supply Chain Risk Score",
        "business_q":     "What is the composite procurement risk level (0-100)?",
        "formula_plain":  "0.4 × vendor_concentration% + 0.3 × maverick_rate% + 0.3 × anomaly_rate%",
        "formula_sql":    "Computed in Python:\nconc   = VENDOR_CONCENTRATION value\nmav    = MAVERICK_BUY_RATE value\nanom   = COUNT(anomaly_count>0)/COUNT(all events) × 100\nrisk   = 0.4×conc + 0.3×mav + 0.3×anom",
        "source_tables":  "Derived from VENDOR_CONCENTRATION, MAVERICK_BUY_RATE, process_mining_events",
        "source_fields":  "process_mining_events.anomaly_count",
        "unit":           "Score (0-100)",
        "target":         "< 30 (Low risk)",
        "rag":            "GREEN <30, AMBER 30-50, RED >50",
        "notes":          "Weights: Vendor Concentration 40%, Maverick Rate 30%, Anomaly Rate 30%.",
    },
    {
        "kpi_code":       "SOD_CONFLICT_COUNT",
        "kpi_name":       "SOD Conflict Count",
        "business_q":     "How many POs were approved by the same person who raised the PR?",
        "formula_plain":  "COUNT(POs WHERE change_log[FRGZU=X].username = pr.requisitioner)",
        "formula_sql":    "SELECT COUNT(DISTINCT po.purchasing_document)\nFROM po_dump po\nJOIN pr_dump pr ON po.purchase_requisition = pr.purchase_requisition\n  AND po.item_of_requisition = pr.item_of_requisition\nJOIN change_log cl ON cl.object_id = po.purchasing_document\n  AND cl.object_class = 'EINKBELEG'\n  AND cl.field_name IN ('FRGZU','FRGKE')\n  AND cl.change_indicator = 'U'\n  AND cl.new_value = 'X'\n  AND cl.username = pr.requisitioner\nWHERE po.purchase_requisition IS NOT NULL",
        "source_tables":  "po_dump, pr_dump, change_log",
        "source_fields":  "pr_dump.requisitioner, change_log.username WHERE field_name='FRGZU' AND new_value='X'",
        "unit":           "Count",
        "target":         "0 (Zero tolerance)",
        "rag":            "GREEN = 0, RED > 0",
        "notes":          "Segregation of Duties: same person raising PR and approving PO bypasses the 4-eyes principle. Each instance requires investigation.",
    },
    {
        "kpi_code":       "PO_GRN_CONVERSION_RATE",
        "kpi_name":       "PO to GRN Conversion Rate (%)",
        "business_q":     "What % of PO lines have received a goods receipt?",
        "formula_plain":  "COUNT(PO lines with GRN) / COUNT(all PO lines) × 100",
        "formula_sql":    "SELECT\n  COUNT(CASE WHEN grn_posting_date IS NOT NULL THEN 1 END) * 100.0\n  / NULLIF(COUNT(*), 0)\nFROM pr_po_grn_invoice\nWHERE purchasing_document IS NOT NULL",
        "source_tables":  "pr_po_grn_invoice",
        "source_fields":  "grn_posting_date",
        "unit":           "%",
        "target":         "> 80%",
        "rag":            "GREEN >80%, AMBER 60-80%, RED <60%",
        "notes":          "Low rate = large number of POs with no delivery confirmation. Exclude service POs (material_group 9905/9906) if needed.",
    },
    {
        "kpi_code":       "DUPLICATE_INVOICE_COUNT",
        "kpi_name":       "Duplicate Invoice Count",
        "business_q":     "How many invoices appear to be duplicates (same vendor + amount + date)?",
        "formula_plain":  "COUNT(*) - COUNT(DISTINCT vendor + amount + posting_date) FROM invoice_dump WHERE document_type IN ('RE','KR')",
        "formula_sql":    "SELECT COUNT(*) - COUNT(DISTINCT vendor || '|' || amount_local_ccy || '|' || posting_date)\nFROM invoice_dump\nWHERE document_type IN ('RE','KR')",
        "source_tables":  "invoice_dump",
        "source_fields":  "vendor, amount_local_ccy, posting_date, document_type",
        "unit":           "Count",
        "target":         "0",
        "rag":            "GREEN = 0, RED > 0",
        "notes":          "Duplicate = same vendor, same amount, same posting date. Each one is a potential double-payment risk.",
    },
    {
        "kpi_code":       "HIGH_VALUE_PO_COUNT",
        "kpi_name":       "High-Value PO Count (Configurable)",
        "business_q":     "How many POs exceed the user-defined high-value threshold?",
        "formula_plain":  "COUNT(DISTINCT purchasing_document) WHERE net_order_value > HIGH_VALUE_PO_THRESHOLD",
        "formula_sql":    "SELECT COUNT(DISTINCT purchasing_document)\nFROM po_dump\nWHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))\n  AND CAST(net_order_value AS REAL) > [kpi_config.HIGH_VALUE_PO_THRESHOLD]",
        "source_tables":  "po_dump, kpi_config",
        "source_fields":  "net_order_value, deletion_indicator",
        "unit":           "Count",
        "target":         "User-defined",
        "rag":            "User-defined",
        "notes":          "User sets threshold in Leadership dashboard header (₹ Cr input). Saves to kpi_config table. Triggers full KPI recompute. Default: ₹1 Cr.",
    },
    {
        "kpi_code":       "SUMMARY_COUNTS",
        "kpi_name":       "P2P Summary Counts",
        "business_q":     "Snapshot counts across all P2P stages",
        "formula_plain":  "Count of: Approved PRs, Approved POs, GRN lines, Invoice lines, Payments, POs without PR",
        "formula_sql":    "Multiple COUNT queries aggregated as JSON",
        "source_tables":  "pr_dump, po_dump, grn_dump, invoice_dump, payment_dump",
        "source_fields":  "Various — see JSON keys",
        "unit":           "JSON",
        "target":         "—",
        "rag":            "POs without PR = RED if > 0",
        "notes":          "JSON keys: approved_pr, approved_po, grn_lines, invoice_lines, payments, po_without_pr",
    },
]

VENDOR_KPIS = [
    {
        "kpi_code":       "ACTIVE_VENDOR_COUNT",
        "kpi_name":       "Active Vendor Count",
        "business_q":     "How many vendors are fully active (all 5 blocks clear)?",
        "formula_plain":  "COUNT(vendors WHERE all 5 block flags are blank/null)",
        "formula_sql":    "SELECT COUNT(DISTINCT vendor)\nFROM vendor_master\nWHERE (central_purchasing_block IS NULL OR central_purchasing_block NOT IN ('X'))\n  AND (central_posting_block    IS NULL OR central_posting_block    NOT IN ('X'))\n  AND (deletion_flag_central    IS NULL OR deletion_flag_central    NOT IN ('X'))\n  AND (payment_block            IS NULL OR payment_block            NOT IN ('*'))\n  AND (posting_block_cc         IS NULL OR posting_block_cc         NOT IN ('X'))",
        "source_tables":  "vendor_master",
        "source_fields":  "central_purchasing_block (SPERR), central_posting_block (SPERM), deletion_flag_central (LOEVM), payment_block (ZAHLS), posting_block_cc (LFB1-SPERR)",
        "unit":           "Count",
        "target":         "Track trend",
        "rag":            "Informational",
        "notes":          "ALL 5 blocks must be clear. payment_block uses '*' not 'X'. Vendor is blocked if ANY one flag is set.",
    },
    {
        "kpi_code":       "VENDOR_BREAKDOWN",
        "kpi_name":       "Vendor Type Breakdown",
        "business_q":     "How many vendors by type: active, blocked, one-time, domestic, international, MSME?",
        "formula_plain":  "Grouped COUNT by vendor_type and block status",
        "formula_sql":    "SELECT SUM(active), SUM(blocked), SUM(one_time),\n       SUM(domestic), SUM(international), SUM(msme), COUNT(*)\nFROM vendor_master (conditional sums)",
        "source_tables":  "vendor_master",
        "source_fields":  "vendor_type (DOMESTIC/INTERNATIONAL/ONE_TIME), msme_flag (M/S), all block flags",
        "unit":           "JSON",
        "target":         "—",
        "rag":            "blocked > 0: RED",
        "notes":          "JSON keys: active, blocked, one_time, domestic, international, msme, total. Vendor types set in vendor_master.vendor_type field.",
    },
    {
        "kpi_code":       "OTIF_RATE",
        "kpi_name":       "OTIF Rate (%)",
        "business_q":     "What % of GRN deliveries were on-time AND in-full?",
        "formula_plain":  "COUNT(GRN WHERE posting_date ≤ expected_delivery_date AND qty ≥ 95% of scheduled) / COUNT(all GRN) × 100",
        "formula_sql":    "SELECT\n  COUNT(CASE\n    WHEN grn.posting_date <= pod.expected_delivery_date\n     AND CAST(grn.quantity AS REAL) >= CAST(pod.scheduled_quantity AS REAL) * 0.95\n    THEN 1 END) * 100.0 / NULLIF(COUNT(*), 0)\nFROM po_delivery_dump pod\nJOIN grn_dump grn ON grn.purchasing_document = pod.purchasing_document\n  AND grn.item = pod.item\nWHERE grn.debit_credit_ind = 'S'",
        "source_tables":  "po_delivery_dump, grn_dump",
        "source_fields":  "grn_dump.posting_date, grn_dump.quantity, grn_dump.debit_credit_ind='S', po_delivery_dump.expected_delivery_date, po_delivery_dump.scheduled_quantity",
        "unit":           "%",
        "target":         "> 90%",
        "rag":            "GREEN >90%, AMBER 75-90%, RED <75%",
        "notes":          "On-Time: GRN posting_date ≤ expected_delivery_date. In-Full: received qty ≥ 95% of scheduled. debit_credit_ind='S' = receipt (not return).",
    },
    {
        "kpi_code":       "AVG_DELIVERY_DELAY",
        "kpi_name":       "Average Delivery Delay (Late Only)",
        "business_q":     "By how many days are late deliveries typically late?",
        "formula_plain":  "AVG(DATEDIFF(grn.posting_date, pod.expected_delivery_date)) WHERE posting_date > expected_delivery_date",
        "formula_sql":    "SELECT AVG(julianday(grn.posting_date) - julianday(pod.expected_delivery_date))\nFROM po_delivery_dump pod\nJOIN grn_dump grn ON grn.purchasing_document = pod.purchasing_document\n  AND grn.item = pod.item\nWHERE grn.debit_credit_ind = 'S'\n  AND grn.posting_date > pod.expected_delivery_date",
        "source_tables":  "po_delivery_dump, grn_dump",
        "source_fields":  "grn_dump.posting_date, po_delivery_dump.expected_delivery_date",
        "unit":           "Days",
        "target":         "≤ 3 days",
        "rag":            "GREEN ≤3, AMBER 3-7, RED >7",
        "notes":          "Only includes late deliveries (posting_date > expected). On-time deliveries excluded from average.",
    },
    {
        "kpi_code":       "QTY_VARIANCE_RATE",
        "kpi_name":       "Quantity Variance Rate (%)",
        "business_q":     "What % of PO lines received less than 95% of ordered quantity (short supply)?",
        "formula_plain":  "COUNT(lines WHERE net_grn_qty < 95% of po.order_quantity) / COUNT(all PO lines) × 100",
        "formula_sql":    "SELECT\n  COUNT(CASE\n    WHEN COALESCE(f.grn_quantity, 0) < CAST(f.po_quantity AS REAL) * 0.95\n    THEN 1 END) * 100.0 / NULLIF(COUNT(*), 0)\nFROM pr_po_grn_invoice f\nWHERE f.purchasing_document IS NOT NULL AND f.po_quantity IS NOT NULL",
        "source_tables":  "pr_po_grn_invoice",
        "source_fields":  "grn_quantity (net S-H), po_quantity",
        "unit":           "%",
        "target":         "< 5%",
        "rag":            "GREEN <5%, AMBER 5-10%, RED >10%",
        "notes":          "grn_quantity is NET (receipts minus returns). Handles multiple GRN postings per PO line.",
    },
    {
        "kpi_code":       "TOP_VENDOR_SPEND",
        "kpi_name":       "Top-10 Vendor Spend Share",
        "business_q":     "Which vendors account for the most spend, and what is their share?",
        "formula_plain":  "For each vendor: SUM(net_order_value) / total_net_order_value × 100, ranked descending, top 10",
        "formula_sql":    "SELECT vendor, vendor_name,\n  SUM(CAST(net_order_value AS REAL)) AS spend,\n  SUM(...) / total * 100 AS share_pct\nFROM po_dump\nGROUP BY vendor\nORDER BY spend DESC LIMIT 10",
        "source_tables":  "po_dump, vendor_master",
        "source_fields":  "vendor, net_order_value, deletion_indicator",
        "unit":           "JSON (%)",
        "target":         "Top vendor < 20% share",
        "rag":            "Any vendor >20%: AMBER; >30%: RED",
        "notes":          "JSON array: [{vendor, name, spend, share_pct}]. Used by Vendor dashboard bar chart.",
    },
    {
        "kpi_code":       "BLOCKED_VENDOR_COUNT",
        "kpi_name":       "Blocked Vendor Count",
        "business_q":     "How many vendors have any purchasing, posting, or payment block active?",
        "formula_plain":  "COUNT(vendors WHERE central_purchasing_block='X' OR central_posting_block='X' OR payment_block='*' OR posting_block_cc='X')",
        "formula_sql":    "SELECT COUNT(DISTINCT vendor) FROM vendor_master\nWHERE central_purchasing_block = 'X'\n   OR central_posting_block    = 'X'\n   OR payment_block            = '*'\n   OR posting_block_cc         = 'X'",
        "source_tables":  "vendor_master",
        "source_fields":  "All 5 block flags",
        "unit":           "Count",
        "target":         "Investigate each",
        "rag":            "GREEN = 0 unexpected blocks, RED > 0",
        "notes":          "Each blocked vendor should be reviewed. Blocked ≠ inactive — vendor may still have open POs which trigger VENDOR_BLOCK anomaly.",
    },
    {
        "kpi_code":       "MSME_VENDOR_COUNT",
        "kpi_name":       "MSME Vendor Count",
        "business_q":     "How many MSME (Micro/Small/Medium Enterprise) vendors are registered?",
        "formula_plain":  "COUNT(*) FROM vendor_master WHERE msme_flag IN ('M','S')",
        "formula_sql":    "SELECT COUNT(*) FROM vendor_master WHERE msme_flag IN ('M','S')",
        "source_tables":  "vendor_master",
        "source_fields":  "msme_flag (M=Micro Enterprise, S=Small Enterprise)",
        "unit":           "Count",
        "target":         "Track trend (regulatory reporting)",
        "rag":            "Informational — regulatory reporting",
        "notes":          "Required for MSMED Act compliance in India. M=Micro (<₹1Cr turnover), S=Small (₹1-10Cr).",
    },
    {
        "kpi_code":       "VENDOR_COMPLIANCE_RATE",
        "kpi_name":       "Vendor Compliance Rate (%)",
        "business_q":     "What % of vendors have no active blocks (fully compliant)?",
        "formula_plain":  "COUNT(vendors with all 5 blocks clear) / COUNT(all vendors) × 100",
        "formula_sql":    "SELECT\n  COUNT(CASE WHEN all_blocks_clear THEN 1 END) * 100.0\n  / NULLIF(COUNT(*), 0)\nFROM vendor_master",
        "source_tables":  "vendor_master",
        "source_fields":  "All 5 block flags",
        "unit":           "%",
        "target":         "> 95%",
        "rag":            "GREEN >95%, AMBER 85-95%, RED <85%",
        "notes":          "Complementary to BLOCKED_VENDOR_COUNT. Used to track vendor risk portfolio health.",
    },
    {
        "kpi_code":       "VENDOR_MASTER_CHANGES",
        "kpi_name":       "Vendor Master Changes (MTD)",
        "business_q":     "How often is vendor master data being modified this month?",
        "formula_plain":  "COUNT(DISTINCT object_id) FROM change_log WHERE object_class='KRED' AND change_date >= month_start",
        "formula_sql":    "SELECT COUNT(DISTINCT object_id)\nFROM change_log\nWHERE object_class = 'KRED'\n  AND change_date >= MTD_START",
        "source_tables":  "change_log",
        "source_fields":  "object_class='KRED', object_id (vendor code), change_date",
        "unit":           "Count",
        "target":         "< 3 per vendor per month",
        "rag":            "HIGH frequency = potential control gap",
        "notes":          "Tracks changes to LFA1 (general) and LFB1 (company code) vendor tables. SAP transaction XK01/XK02.",
    },
]

UTILIZATION_KPIS = [
    {
        "kpi_code":       "IT_SPEND_YTD",
        "kpi_name":       "IT/Software Spend (YTD)",
        "business_q":     "How much has been spent on IT tools and software this FY?",
        "formula_plain":  "SUM(net_order_value) WHERE material_group IN ('IT','CLOUD','LICENSE','SOFTWARE','SAAS','9904','9905') AND document_date >= FY_start",
        "formula_sql":    "SELECT SUM(CAST(net_order_value AS REAL))\nFROM po_dump\nWHERE (deletion_indicator IS NULL OR deletion_indicator NOT IN ('L','X'))\n  AND document_date >= '{FY_YEAR}-04-01'\n  AND UPPER(material_group) IN ('IT','CLOUD','LICENSE','SOFTWARE','SAAS','9904','9905')",
        "source_tables":  "po_dump",
        "source_fields":  "net_order_value, material_group, document_date, deletion_indicator",
        "unit":           "INR",
        "target":         "Within IT budget",
        "rag":            "vs IT budget allocation",
        "notes":          "Material groups 9904=IT Hardware, 9905=IT Software/Licenses. Also matches text codes 'IT','CLOUD','SOFTWARE','SAAS'.",
    },
    {
        "kpi_code":       "CAPEX_OPEX_SPLIT",
        "kpi_name":       "CAPEX vs OPEX Split",
        "business_q":     "What proportion of spend is capital (CAPEX) vs operational (OPEX)?",
        "formula_plain":  "SUM(net_order_value) WHERE capex_opex_flag='CAPEX' / total and similar for OPEX",
        "formula_sql":    "SELECT\n  SUM(CASE WHEN UPPER(capex_opex_flag)='CAPEX' THEN CAST(net_order_value AS REAL) ELSE 0 END),\n  SUM(CASE WHEN UPPER(capex_opex_flag)='OPEX'  THEN CAST(net_order_value AS REAL) ELSE 0 END)\nFROM po_dump\nWHERE document_date >= FY_START AND deletion_indicator NOT IN ('L','X')",
        "source_tables":  "po_dump",
        "source_fields":  "capex_opex_flag (set from material_group mapping), net_order_value",
        "unit":           "JSON (INR + %)",
        "target":         "Managed per budget policy",
        "rag":            "vs approved CAPEX/OPEX budget split",
        "notes":          "capex_opex_flag derived from material_group: 9902/9904/9905=CAPEX, all others=OPEX. JSON: {capex, opex, capex_pct, opex_pct}.",
    },
    {
        "kpi_code":       "AVG_LICENSE_UTILIZATION",
        "kpi_name":       "Avg License Utilization Rate (%)",
        "business_q":     "What % of purchased licenses are actively being used?",
        "formula_plain":  "AVG(active_users / total_licenses × 100) per tool FROM license_usage",
        "formula_sql":    "SELECT AVG(CAST(active_users AS REAL) * 100.0 / CAST(total_licenses AS REAL))\nFROM license_usage WHERE total_licenses > 0",
        "source_tables":  "license_usage (separate upload — not from SAP)",
        "source_fields":  "active_users, total_licenses",
        "unit":           "%",
        "target":         "> 80%",
        "rag":            "GREEN >80%, AMBER 50-80%, RED <50%",
        "notes":          "license_usage table fed from IT SAM tools (Flexera, Snow, ServiceNow). Separate from PO data. Upload via admin interface.",
    },
    {
        "kpi_code":       "COST_PER_USER",
        "kpi_name":       "Avg Cost Per Active User",
        "business_q":     "How much are we spending per active user across all licensed tools?",
        "formula_plain":  "AVG(annual_cost_inr / active_users) per tool FROM license_usage WHERE active_users > 0",
        "formula_sql":    "SELECT AVG(CAST(annual_cost_inr AS REAL) / NULLIF(CAST(active_users AS REAL), 0))\nFROM license_usage WHERE active_users > 0",
        "source_tables":  "license_usage",
        "source_fields":  "annual_cost_inr, active_users",
        "unit":           "INR / user",
        "target":         "Benchmarked per tool category",
        "rag":            "Compare to market benchmarks",
        "notes":          "Annual cost per active user. High cost + low utilization = renegotiation target.",
    },
    {
        "kpi_code":       "UNDERUTILIZED_LICENSE_COUNT",
        "kpi_name":       "Underutilized License Count",
        "business_q":     "How many tools have less than 50% of licenses actively used?",
        "formula_plain":  "COUNT(tools WHERE active_users / total_licenses < 50%)",
        "formula_sql":    "SELECT COUNT(*) FROM license_usage\nWHERE CAST(active_users AS REAL) * 1.0 / CAST(total_licenses AS REAL) < 0.50",
        "source_tables":  "license_usage",
        "source_fields":  "active_users, total_licenses",
        "unit":           "Count",
        "target":         "≤ 2 tools",
        "rag":            "GREEN ≤2, RED >3",
        "notes":          "Tools below 50% utilization are candidates for contract renegotiation or downsizing.",
    },
    {
        "kpi_code":       "UPCOMING_RENEWALS",
        "kpi_name":       "Upcoming License Renewals (60 days)",
        "business_q":     "Which IT POs/contracts are renewing in the next 60 days?",
        "formula_plain":  "COUNT(POs WHERE material_group=IT category AND expected_delivery_date BETWEEN today AND today+60)",
        "formula_sql":    "SELECT COUNT(DISTINCT po.purchasing_document)\nFROM po_dump po\nLEFT JOIN po_delivery_dump pod ON pod.purchasing_document = po.purchasing_document\n  AND pod.item = po.item\nWHERE UPPER(po.material_group) IN ('IT','CLOUD','LICENSE','SOFTWARE','SAAS','9904','9905')\n  AND COALESCE(pod.expected_delivery_date, po.delivery_date)\n    BETWEEN date('now') AND date('now', '+60 days')",
        "source_tables":  "po_dump, po_delivery_dump",
        "source_fields":  "material_group, po_delivery_dump.expected_delivery_date",
        "unit":           "Count",
        "target":         "All tracked (procurement action required)",
        "rag":            "Informational — triggers review",
        "notes":          "60-day window gives time for renegotiation, downsize decision, or renewal action.",
    },
    {
        "kpi_code":       "WASTED_LICENSE_COST_MONTHLY",
        "kpi_name":       "Monthly Wasted License Cost",
        "business_q":     "How much are we wasting monthly on unused license capacity?",
        "formula_plain":  "SUM((total_licenses - CEILING(active_users / 0.80)) × unit_monthly_cost) WHERE utilization < 80%",
        "formula_sql":    "SELECT SUM(\n  (total_licenses - CAST(CEIL(CAST(active_users AS REAL) / 0.80) AS INTEGER))\n  * (annual_cost_inr / 12.0 / NULLIF(total_licenses, 0))\n)\nFROM license_usage\nWHERE CAST(active_users AS REAL) / CAST(total_licenses AS REAL) < 0.80\n  AND total_licenses > 0",
        "source_tables":  "license_usage",
        "source_fields":  "total_licenses, active_users, annual_cost_inr",
        "unit":           "INR / month",
        "target":         "Minimize",
        "rag":            "Higher = more savings potential",
        "notes":          "Formula: ideal_licenses = CEILING(active_users / 0.80) i.e. rightsized to 80% utilization. Excess = total - ideal. unit_monthly_cost = annual_cost / 12 / total_licenses.",
    },
]

ANOMALY_RULES = [
    {
        "anomaly_code":  "MAVERICK_BUY",
        "severity":      "HIGH",
        "description":   "PO created without upstream PR (off-process buying)",
        "detection":     "pr_po_grn_invoice.is_maverick = 1 (po.purchase_requisition IS NULL or '')",
        "business_risk": "No budget approval or requirement validation. Bypasses controls.",
        "sap_field":     "EKPO-BANFN (purchase_requisition) NULL",
    },
    {
        "anomaly_code":  "LATE_DELIVERY",
        "severity":      "MEDIUM",
        "description":   "GRN posting date exceeds expected delivery date (po_delivery_dump)",
        "detection":     "grn.posting_date > po_delivery_dump.expected_delivery_date",
        "business_risk": "Production delays, SLA breach, potential penalty clauses.",
        "sap_field":     "EKBE-BUDAT vs EKET-EINDT",
    },
    {
        "anomaly_code":  "THREE_WAY_MISMATCH",
        "severity":      "HIGH",
        "description":   "Invoice quantity differs from net GRN quantity by > 5%",
        "detection":     "|grn_quantity (net) - invoice_quantity| / invoice_quantity > 0.05",
        "business_risk": "Paying for undelivered goods. Over/under-invoicing.",
        "sap_field":     "EKBE-MENGE (VGABE=E vs Q)",
    },
    {
        "anomaly_code":  "DUPLICATE_INVOICE",
        "severity":      "HIGH",
        "description":   "Same vendor + amount + posting_date appears more than once in invoice_dump",
        "detection":     "GROUP BY vendor, amount_local_ccy, posting_date HAVING COUNT(*) > 1",
        "business_risk": "Double payment risk. Vendor may extract duplicate funds.",
        "sap_field":     "BSIK/BSAK duplicate BELNR pattern",
    },
    {
        "anomaly_code":  "BACKDATED_PO",
        "severity":      "MEDIUM",
        "description":   "PO document_date is earlier than PR release_date",
        "detection":     "po.document_date < pr.release_date (item-level join)",
        "business_risk": "PO created before PR was approved — retrospective approval (rubber-stamping).",
        "sap_field":     "EKKO-BEDAT vs EBAN-FRGDT",
    },
    {
        "anomaly_code":  "PAYMENT_BEFORE_GRN",
        "severity":      "HIGH",
        "description":   "Payment clearing date is before GRN posting date",
        "detection":     "payment.posting_date < grn_posting_date (via invoice chain)",
        "business_risk": "Paying before goods received. Fraud risk. No recourse if goods don't arrive.",
        "sap_field":     "BSAK-AUGDT vs EKBE-BUDAT (VGABE=E)",
    },
    {
        "anomaly_code":  "PRICE_DEVIATION",
        "severity":      "MEDIUM",
        "description":   "PO net_order_price deviates > 10% from PR valuation_price",
        "detection":     "|po_net_price - pr_value| / pr_value > 0.10",
        "business_risk": "Price inflation after PR approval. Contract not honored.",
        "sap_field":     "EKPO-NETPR vs EBAN-PREIS",
    },
    {
        "anomaly_code":  "SPLIT_PO",
        "severity":      "HIGH",
        "description":   "Multiple POs to same vendor on same date with same material_group — each below approval threshold but total above",
        "detection":     "GROUP BY vendor, document_date, material_group HAVING COUNT(*) > 1",
        "business_risk": "Deliberate splitting to bypass DOA (Delegation of Authority) approval levels.",
        "sap_field":     "EKKO-EBELN grouped by LIFNR + BEDAT + EKPO-MATKL",
    },
    {
        "anomaly_code":  "VENDOR_BLOCK",
        "severity":      "HIGH",
        "description":   "PO raised against a vendor with central_purchasing_block='X'",
        "detection":     "po.vendor EXISTS IN vendor_master WHERE central_purchasing_block='X'",
        "business_risk": "Procurement from sanctioned/suspended/problematic vendor. Compliance violation.",
        "sap_field":     "LFA1-SPERR = 'X'",
    },
    {
        "anomaly_code":  "SOD_VIOLATION",
        "severity":      "HIGH",
        "description":   "PR requisitioner = PO approver (same person raised PR and approved PO)",
        "detection":     "change_log[FRGZU/FRGKE=X].username = pr_dump.requisitioner for same PR→PO chain",
        "business_risk": "Bypasses 4-eyes / dual control principle. Single person controls full buy cycle.",
        "sap_field":     "EBAN-AFNAM vs CDHDR-USERNAME WHERE CDPOS-FNAME='FRGZU' AND NEW='X'",
    },
]

DUE_DATE_LOGIC = {
    "title":         "Invoice Due Date Calculation",
    "formula":       "due_date = baseline_date + days_1",
    "baseline_date": "SAP field ZFBDT — the contractual reference date for payment terms. Typically = vendor invoice date (BLDAT) unless manually overridden.",
    "days_1":        "SAP field ZBD3T — net payment days from payment terms master (T052U). Examples: N030=30 days, N045=45 days, N060=60 days.",
    "standard_terms": {
        "N030": 30, "N045": 45, "N060": 60, "I001": 90
    },
    "implementation": "parser.py _compute_due_date() fills due_date = pd.to_datetime(baseline_date) + timedelta(days=days_1) if due_date is null/empty.",
    "storage":       "invoice_dump.baseline_date (TEXT YYYY-MM-DD), invoice_dump.days_1 (TEXT integer), invoice_dump.due_date (TEXT YYYY-MM-DD)"
}

COMPOSITE_KEYS = [
    {"entity": "PR Line",       "key_format": "PR|{purchase_requisition}|{item_of_requisition}",        "example": "PR|10003478|00010", "purpose": "Unique PR line identifier for item-level joins"},
    {"entity": "PO Line",       "key_format": "PO|{purchasing_document}|{item}",                         "example": "PO|2000001883|00010", "purpose": "Unique PO line identifier"},
    {"entity": "Entity Filter", "key_format": "{company_code}|{purchasing_org}|{plant}",                 "example": "1001|1000|SDPL", "purpose": "Company/Org/Plant cascade filter for multi-entity reporting"},
    {"entity": "GRN Key",       "key_format": "{material_document}|{material_doc_item}",                 "example": "5000835597|0001", "purpose": "Unique GRN line (natural key)"},
    {"entity": "Invoice Key",   "key_format": "{invoice_doc}|{invoice_year}",                            "example": "5105418119|2022", "purpose": "Unique invoice across fiscal years"},
    {"entity": "Payment Key",   "key_format": "{payment_doc}|{payment_year}",                            "example": "2900021526|2022", "purpose": "Unique payment across fiscal years"},
    {"entity": "Change Key",    "key_format": "{change_number}|{table_name}|{field_name}",               "example": "840264|LFA1|ZTERM", "purpose": "Unique change log entry"},
]

PLANT_HIERARCHY = [
    {"company_code": "1001", "company_name": "IntelliSource India Pvt Ltd", "purchasing_org": "1000", "plant": "SDPL", "plant_name": "South Delhi Plant"},
    {"company_code": "1001", "company_name": "IntelliSource India Pvt Ltd", "purchasing_org": "1000", "plant": "BLRP", "plant_name": "Bengaluru Plant"},
    {"company_code": "1001", "company_name": "IntelliSource India Pvt Ltd", "purchasing_org": "2000", "plant": "DELP", "plant_name": "Delhi North Plant"},
    {"company_code": "1001", "company_name": "IntelliSource India Pvt Ltd", "purchasing_org": "2000", "plant": "HYDP", "plant_name": "Hyderabad Plant"},
    {"company_code": "1001", "company_name": "IntelliSource India Pvt Ltd", "purchasing_org": "3000", "plant": "MNAL", "plant_name": "Mumbai Plant"},
]

# ── WORKBOOK BUILD ─────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ── HELPER: Build a KPI sheet ─────────────────────────────────────────────────

def build_kpi_sheet(wb, sheet_name, kpi_list, color_key, target_user, refresh, purpose):
    main_color, light_color, title_text = DASHBOARD_COLORS[color_key]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = main_color

    # Col widths
    widths = {"A": 5, "B": 8, "C": 30, "D": 38, "E": 50, "F": 65, "G": 38, "H": 30, "I": 20, "J": 14, "K": 50}
    for col, w in widths.items():
        set_col_width(ws, col, w)

    # Row 1 — Title banner
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = f"  {title_text}"
    c.fill = fill(main_color)
    c.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    c.alignment = align("left", "center")
    ws.row_dimensions[1].height = 32

    # Row 2 — Subtitle
    ws.merge_cells("A2:K2")
    c = ws["A2"]
    c.value = f"  Target: {target_user}  |  Refresh: {refresh}  |  {purpose}"
    c.fill = fill(light_color)
    c.font = Font(bold=False, color="000000", size=9, name="Calibri")
    c.alignment = align("left", "center")
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 6  # spacer

    # Row 4 — Column headers
    headers = ["#", "KPI Code", "KPI Name", "Business Question",
               "Formula (Plain English)", "Formula (SQL / Logic)",
               "Source Tables", "Source Fields", "Unit", "Target / Threshold", "Notes"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.fill = fill(main_color)
        c.font = font(bold=True, color=WHITE, size=10)
        c.border = border()
        c.alignment = align("center", "center", wrap=True)
    ws.row_dimensions[4].height = 28

    # Data rows
    for i, kpi in enumerate(kpi_list):
        row = i + 5
        vals = [
            i + 1,
            kpi["kpi_code"],
            kpi["kpi_name"],
            kpi["business_q"],
            kpi["formula_plain"],
            kpi["formula_sql"],
            kpi["source_tables"],
            kpi["source_fields"],
            kpi["unit"],
            kpi["target"],
            kpi["notes"],
        ]
        for col_idx, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col_idx, value=v)
            bg = light_color if i % 2 == 0 else WHITE
            c.fill = fill(bg)
            c.font = Font(color="000000", size=9, name="Calibri")
            c.border = border()
            c.alignment = align("left", "center", wrap=True)
            if col_idx in (2,):  # KPI code — bold mono
                c.font = Font(color=main_color, size=9, bold=True, name="Courier New")
            if col_idx in (6,):  # SQL — smaller mono
                c.font = Font(color="333333", size=8, name="Courier New")
                c.alignment = align("left", "top", wrap=True)
        ws.row_dimensions[row].height = 64

    # Freeze panes
    ws.freeze_panes = "C5"
    return ws

# ── Sheet 1: Summary / Index ───────────────────────────────────────────────────

ws_sum = wb.create_sheet("📋 Summary")
ws_sum.sheet_view.showGridLines = False
ws_sum.sheet_properties.tabColor = KPMG_BLUE

for col, w in {"A": 4, "B": 32, "C": 20, "D": 12, "E": 12, "F": 50}.items():
    set_col_width(ws_sum, col, w)

# Title
ws_sum.merge_cells("A1:F1")
c = ws_sum["A1"]
c.value = "  IntelliSource P2P — KPI & Anomaly Documentation"
c.fill = fill(KPMG_BLUE)
c.font = Font(bold=True, color=WHITE, size=16, name="Calibri")
c.alignment = align("left", "center")
ws_sum.row_dimensions[1].height = 40

ws_sum.merge_cells("A2:F2")
c = ws_sum["A2"]
c.value = f"  Generated by IntelliSource P2P | {len(PROCUREMENT_KPIS)+len(FINANCIAL_KPIS)+len(LEADERSHIP_KPIS)+len(VENDOR_KPIS)+len(UTILIZATION_KPIS)} KPIs across 5 dashboards | {len(ANOMALY_RULES)} Anomaly Rules"
c.fill = fill(KPMG_BLUE_LIGHT)
c.font = Font(bold=False, color="000000", size=10, name="Calibri")
c.alignment = align("left", "center")
ws_sum.row_dimensions[2].height = 20

ws_sum.row_dimensions[3].height = 8

# Dashboard summary table
dash_headers = ["#", "Dashboard", "Sheet", "# KPIs", "Target User", "Refresh"]
for ci, h in enumerate(dash_headers, 1):
    c = ws_sum.cell(row=4, column=ci, value=h)
    c.fill = fill(KPMG_BLUE)
    c.font = font(bold=True, size=10)
    c.border = border()
    c.alignment = align("center", "center")
ws_sum.row_dimensions[4].height = 22

dash_data = [
    (1, "Procurement",  "01 Procurement",  len(PROCUREMENT_KPIS), "Procurement Managers", "Weekly"),
    (2, "Financial",    "02 Financial",     len(FINANCIAL_KPIS),   "Finance Team",          "Monthly"),
    (3, "Leadership",   "03 Leadership",    len(LEADERSHIP_KPIS),  "CXO / Leadership",      "Monthly"),
    (4, "Vendor",       "04 Vendor",        len(VENDOR_KPIS),      "Procurement Managers",  "Monthly"),
    (5, "Utilization",  "05 Utilization",   len(UTILIZATION_KPIS), "IT / Resource Managers","Monthly"),
]

row_colors = [KPMG_BLUE_LIGHT, KPMG_TEAL_LIGHT, KPMG_PURP_LIGHT, KPMG_CYAN_LIGHT, "D6EDDE"]
for i, (num, dash, sheet, count, user, refresh) in enumerate(dash_data):
    r = i + 5
    for ci, v in enumerate([num, dash, sheet, count, user, refresh], 1):
        c = ws_sum.cell(row=r, column=ci, value=v)
        c.fill = fill(row_colors[i])
        c.font = Font(color="000000", size=9, name="Calibri")
        c.border = border()
        c.alignment = align("center" if ci in (1,4) else "left", "center")
    ws_sum.row_dimensions[r].height = 18

ws_sum.row_dimensions[11].height = 12

# KPI Count summary
ws_sum.cell(row=12, column=1, value="Total KPIs:").font = Font(bold=True, size=10, name="Calibri")
ws_sum.cell(row=12, column=2, value=len(PROCUREMENT_KPIS)+len(FINANCIAL_KPIS)+len(LEADERSHIP_KPIS)+len(VENDOR_KPIS)+len(UTILIZATION_KPIS)).font = Font(bold=True, color=KPMG_BLUE, size=12)
ws_sum.cell(row=13, column=1, value="Anomaly Rules:").font = Font(bold=True, size=10, name="Calibri")
ws_sum.cell(row=13, column=2, value=len(ANOMALY_RULES)).font = Font(bold=True, color=RED, size=12)

ws_sum.row_dimensions[15].height = 12

# Due date note
ws_sum.merge_cells("A16:F16")
c = ws_sum["A16"]
c.value = "  Due Date Formula: due_date = baseline_date (ZFBDT) + days_1 (ZBD3T)"
c.fill = fill(AMBER_LIGHT)
c.font = Font(bold=True, color="000000", size=10, name="Calibri")
c.border = border()
c.alignment = align("left", "center")
ws_sum.row_dimensions[16].height = 20

ws_sum.merge_cells("A17:F17")
c = ws_sum["A17"]
c.value = "  High-Value PO Threshold: Configurable from Leadership dashboard header (PUT /api/kpi-config/HIGH_VALUE_PO_THRESHOLD). Default: ₹1 Cr."
c.fill = fill(AMBER_LIGHT)
c.font = Font(bold=False, color="000000", size=9, name="Calibri")
c.border = border()
c.alignment = align("left", "center")
ws_sum.row_dimensions[17].height = 18


# ── Build KPI Sheets ──────────────────────────────────────────────────────────

build_kpi_sheet(wb, "01 Procurement", PROCUREMENT_KPIS, "Procurement",
                "Procurement Managers", "Weekly",
                "Real-time visibility into PO activity, approvals, and process health")

build_kpi_sheet(wb, "02 Financial", FINANCIAL_KPIS, "Financial",
                "Finance Team", "Monthly",
                "Spend vs budget tracking, invoice/payment cycle health")

build_kpi_sheet(wb, "03 Leadership", LEADERSHIP_KPIS, "Leadership",
                "CXO / Leadership", "Monthly",
                "Strategic portfolio view — scannable in 30 seconds")

build_kpi_sheet(wb, "04 Vendor", VENDOR_KPIS, "Vendor",
                "Procurement Managers", "Monthly",
                "Vendor performance, compliance, and concentration risk")

build_kpi_sheet(wb, "05 Utilization", UTILIZATION_KPIS, "Utilization",
                "IT / Resource Managers", "Monthly",
                "IT/software license utilization and CAPEX/OPEX spend")


# ── Sheet: Anomaly Rules ──────────────────────────────────────────────────────

ws_anom = wb.create_sheet("06 Anomaly Rules")
ws_anom.sheet_view.showGridLines = False
ws_anom.sheet_properties.tabColor = RED

for col, w in {"A": 5, "B": 30, "C": 12, "D": 45, "E": 55, "F": 45, "G": 40}.items():
    set_col_width(ws_anom, col, w)

ws_anom.merge_cells("A1:G1")
c = ws_anom["A1"]
c.value = "  ANOMALY DETECTION RULES — P2P Process Mining"
c.fill = fill(RED)
c.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
c.alignment = align("left", "center")
ws_anom.row_dimensions[1].height = 32

ws_anom.merge_cells("A2:G2")
c = ws_anom["A2"]
c.value = f"  {len(ANOMALY_RULES)} rules run after every ETL upload. Results stored in process_mining_events.anomaly_flags (comma-separated)."
c.fill = fill(RED_LIGHT)
c.font = Font(bold=False, color="000000", size=9, name="Calibri")
c.alignment = align("left", "center")
ws_anom.row_dimensions[2].height = 16
ws_anom.row_dimensions[3].height = 6

headers = ["#", "Anomaly Code", "Severity", "Description", "Detection Logic", "Business Risk", "SAP Source Field"]
for ci, h in enumerate(headers, 1):
    c = ws_anom.cell(row=4, column=ci, value=h)
    c.fill = fill(RED)
    c.font = font(bold=True, size=10)
    c.border = border()
    c.alignment = align("center", "center", wrap=True)
ws_anom.row_dimensions[4].height = 24

sev_colors = {"HIGH": RED_LIGHT, "MEDIUM": AMBER_LIGHT, "LOW": GREEN_LIGHT}

for i, anom in enumerate(ANOMALY_RULES):
    r = i + 5
    vals = [i+1, anom["anomaly_code"], anom["severity"],
            anom["description"], anom["detection"], anom["business_risk"], anom["sap_field"]]
    for ci, v in enumerate(vals, 1):
        c = ws_anom.cell(row=r, column=ci, value=v)
        bg = sev_colors.get(anom["severity"], WHITE)
        c.fill = fill(bg)
        c.font = Font(color="000000", size=9, name="Calibri")
        c.border = border()
        c.alignment = align("left", "center", wrap=True)
        if ci == 2:
            c.font = Font(color=RED, bold=True, size=9, name="Courier New")
        if ci == 3:
            sev = anom["severity"]
            c.font = Font(color=RED if sev=="HIGH" else (AMBER if sev=="MEDIUM" else GREEN),
                          bold=True, size=9, name="Calibri")
            c.alignment = align("center", "center")
    ws_anom.row_dimensions[r].height = 48

ws_anom.freeze_panes = "C5"


# ── Sheet: Due Date & Payment Matching ───────────────────────────────────────

ws_inv = wb.create_sheet("07 Due Date & Payment")
ws_inv.sheet_view.showGridLines = False
ws_inv.sheet_properties.tabColor = KPMG_TEAL

for col, w in {"A": 5, "B": 30, "C": 55, "D": 30}.items():
    set_col_width(ws_inv, col, w)

ws_inv.merge_cells("A1:D1")
c = ws_inv["A1"]
c.value = "  INVOICE DUE DATE FORMULA & PAYMENT MATCHING LOGIC"
c.fill = fill(KPMG_TEAL)
c.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
c.alignment = align("left", "center")
ws_inv.row_dimensions[1].height = 30

row = 3
def add_section(ws, r, title, color=KPMG_TEAL_LIGHT):
    ws.merge_cells(f"A{r}:D{r}")
    c = ws.cell(row=r, column=1, value=f"  {title}")
    c.fill = fill(color)
    c.font = Font(bold=True, color="000000", size=10, name="Calibri")
    c.alignment = align("left", "center")
    ws.row_dimensions[r].height = 20
    return r + 1

def add_row(ws, r, label, value):
    c1 = ws.cell(row=r, column=2, value=label)
    c1.font = Font(bold=True, size=9, name="Calibri")
    c1.border = border()
    c1.fill = fill(LIGHT_GREY)
    c1.alignment = align("right", "center")
    ws.merge_cells(f"C{r}:D{r}")
    c2 = ws.cell(row=r, column=3, value=value)
    c2.font = Font(size=9, name="Calibri")
    c2.border = border()
    c2.fill = fill(WHITE)
    c2.alignment = align("left", "center", wrap=True)
    ws.row_dimensions[r].height = 30
    return r + 1

row = add_section(ws_inv, row, "Due Date Formula")
row = add_row(ws_inv, row, "Formula", "due_date = baseline_date + days_1")
row = add_row(ws_inv, row, "baseline_date (ZFBDT)", "The SAP contractual reference date from which payment terms count. Typically = vendor_invoice_date (BLDAT). Can be manually set by AP team.")
row = add_row(ws_inv, row, "days_1 (ZBD3T)", "Net payment days from SAP payment terms master (T052U). Examples: N030=30 days, N045=45 days, N060=60 days, I001=90 days.")
row = add_row(ws_inv, row, "Source fields", "invoice_dump.baseline_date + invoice_dump.days_1 → invoice_dump.due_date")
row = add_row(ws_inv, row, "Implementation", "parser.py _compute_due_date() fills due_date if blank: pd.to_datetime(baseline_date) + timedelta(days=int(days_1))")
row += 1

# Payment terms table
row = add_section(ws_inv, row, "Standard Payment Terms")
terms_data = [("N030", "30", "Net 30 days from baseline"), ("N045", "45", "Net 45 days from baseline"),
              ("N060", "60", "Net 60 days from baseline"), ("I001", "90", "Custom 90-day term")]
for code, days, desc in terms_data:
    ws_inv.cell(row=row, column=2, value=code).font = Font(bold=True, size=9, color=KPMG_TEAL)
    ws_inv.cell(row=row, column=3, value=f"+{days} days").font = Font(size=9)
    ws_inv.cell(row=row, column=4, value=desc).font = Font(size=9, color=DARK_GREY)
    for c in [2,3,4]:
        ws_inv.cell(row=row, column=c).border = border()
        ws_inv.cell(row=row, column=c).fill = fill(KPMG_TEAL_LIGHT if row%2==0 else WHITE)
    ws_inv.row_dimensions[row].height = 18
    row += 1

row += 1
row = add_section(ws_inv, row, "Payment vs Invoice Matching Logic")
row = add_row(ws_inv, row, "Primary link", "invoice_dump.clearing_doc = payment_dump.payment_doc (SAP: BSAK-AUGBL = BSAK-BELNR)")
row = add_row(ws_inv, row, "Reverse link", "payment_dump.cleared_invoice = invoice_dump.invoice_doc (SAP: BSAK-REBZG)")
row = add_row(ws_inv, row, "Year safety", "Match also uses invoice_year to prevent cross-year collisions (invoice_doc not globally unique across years)")
row = add_row(ws_inv, row, "PO invoice chain", "po_invoice_dump.invoice_doc + invoice_year → invoice_dump.invoice_doc + invoice_year → payment_dump.cleared_invoice")
row = add_row(ws_inv, row, "Partial payments", "One invoice may be cleared by multiple payments. Use SUM(payment.amount) per invoice_doc for total paid.")
row = add_row(ws_inv, row, "On-Time check",   "ON_TIME_PAYMENT_RATE: payment.posting_date ≤ invoice.due_date (not invoice.posting_date)")
row = add_row(ws_inv, row, "DPO baseline",    "DPO uses invoice.posting_date as baseline (not due_date). DPO = AVG(payment.posting_date - invoice.posting_date)")


# ── Sheet: Composite Keys & Entity Hierarchy ──────────────────────────────────

ws_keys = wb.create_sheet("08 Composite Keys & Hierarchy")
ws_keys.sheet_view.showGridLines = False
ws_keys.sheet_properties.tabColor = KPMG_PURPLE

for col, w in {"A": 5, "B": 22, "C": 45, "D": 35, "E": 45}.items():
    set_col_width(ws_keys, col, w)

ws_keys.merge_cells("A1:E1")
c = ws_keys["A1"]
c.value = "  COMPOSITE KEYS & COMPANY / PLANT HIERARCHY"
c.fill = fill(KPMG_PURPLE)
c.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
c.alignment = align("left", "center")
ws_keys.row_dimensions[1].height = 30

# Composite keys table
row = 3
ws_keys.merge_cells(f"A{row}:E{row}")
c = ws_keys.cell(row=row, column=1, value="  COMPOSITE KEYS — Line-Level Tracking")
c.fill = fill(KPMG_PURP_LIGHT)
c.font = Font(bold=True, size=10)
c.alignment = align("left", "center")
ws_keys.row_dimensions[row].height = 20
row += 1

headers = ["#", "Entity", "Key Format", "Example", "Purpose"]
for ci, h in enumerate(headers, 1):
    c = ws_keys.cell(row=row, column=ci, value=h)
    c.fill = fill(KPMG_PURPLE)
    c.font = font(bold=True, size=10)
    c.border = border()
    c.alignment = align("center", "center")
ws_keys.row_dimensions[row].height = 22
row += 1

for i, key in enumerate(COMPOSITE_KEYS):
    vals = [i+1, key["entity"], key["key_format"], key["example"], key["purpose"]]
    for ci, v in enumerate(vals, 1):
        c = ws_keys.cell(row=row, column=ci, value=v)
        c.fill = fill(KPMG_PURP_LIGHT if i%2==0 else WHITE)
        is_code_col = ci in (3, 4)
        c.font = Font(size=9,
                      name="Courier New" if is_code_col else "Calibri",
                      color=KPMG_PURPLE if is_code_col else "000000")
        c.border = border()
        c.alignment = align("left", "center", wrap=True)
    ws_keys.row_dimensions[row].height = 22
    row += 1

row += 1
# Plant hierarchy table
ws_keys.merge_cells(f"A{row}:E{row}")
c = ws_keys.cell(row=row, column=1, value="  COMPANY / ORG / PLANT HIERARCHY")
c.fill = fill(KPMG_PURP_LIGHT)
c.font = Font(bold=True, size=10)
c.alignment = align("left", "center")
ws_keys.row_dimensions[row].height = 20
row += 1

headers = ["company_code", "company_name", "purchasing_org", "plant", "plant_name"]
for ci, h in enumerate(headers, 1):
    c = ws_keys.cell(row=row, column=ci, value=h)
    c.fill = fill(KPMG_PURPLE)
    c.font = font(bold=True, size=10)
    c.border = border()
    c.alignment = align("center", "center")
ws_keys.row_dimensions[row].height = 22
row += 1

for i, ph in enumerate(PLANT_HIERARCHY):
    vals = [ph["company_code"], ph["company_name"], ph["purchasing_org"], ph["plant"], ph["plant_name"]]
    for ci, v in enumerate(vals, 1):
        c = ws_keys.cell(row=row, column=ci, value=v)
        c.fill = fill(KPMG_PURP_LIGHT if i%2==0 else WHITE)
        c.font = Font(size=9, name="Calibri")
        c.border = border()
        c.alignment = align("center" if ci in (1,3,4) else "left", "center")
    ws_keys.row_dimensions[row].height = 18
    row += 1

row += 1
ws_keys.merge_cells(f"A{row}:E{row}")
c = ws_keys.cell(row=row, column=1,
    value="  Filter cascade: Company (1001) → Purchasing Org (1000/2000/3000) → Plant (SDPL/BLRP/DELP/HYDP/MNAL)  |  PO entity_key = company_code + '|' + purchasing_org + '|' + plant")
c.fill = fill(AMBER_LIGHT)
c.font = Font(size=9, italic=True)
c.alignment = align("left", "center", wrap=True)
c.border = border()
ws_keys.row_dimensions[row].height = 30


# ── Save ──────────────────────────────────────────────────────────────────────

OUT_PATH = Path(__file__).parent / "IntelliSource_KPI_Documentation.xlsx"
wb.save(str(OUT_PATH))
print(f"\nSaved: {OUT_PATH}")
print(f"  Sheets: {', '.join([ws.title for ws in wb.worksheets])}")
print(f"  Total KPIs: {len(PROCUREMENT_KPIS)+len(FINANCIAL_KPIS)+len(LEADERSHIP_KPIS)+len(VENDOR_KPIS)+len(UTILIZATION_KPIS)}")
print(f"  Anomaly rules: {len(ANOMALY_RULES)}")
